import io
import json
import os
import re
import sys
import threading
import time
import urllib.parse
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
import requests

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference


def _configure_stdio() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is None:
            continue
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(errors="replace")
            except Exception:
                pass


_configure_stdio()


BASE_DIR = Path(__file__).resolve().parent
LIVE_DIR = BASE_DIR.parent
REPO_ROOT = LIVE_DIR.parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# B 站爬虫位于 <repo>/_2_bilibili/live/
_BILI_LIVE_DIR = REPO_ROOT / "_2_bilibili" / "live"
for _p in (str(LIVE_DIR), str(_BILI_LIVE_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import search_bilibili_live_rooms as bili_crawler  # noqa: E402

# 抖音爬虫位于 <repo>/快手抖音B站直播/live/
DOUYIN_LIVE_DIR = LIVE_DIR / "live"
if str(DOUYIN_LIVE_DIR) not in sys.path:
    sys.path.insert(0, str(DOUYIN_LIVE_DIR))
import douyin_search_live_rooms as douyin_crawler  # noqa: E402

DOUYIN_COOKIE_FILE = DOUYIN_LIVE_DIR / "douyin_cookie.txt"
DOUYIN_COOKIE_AUTO_REFRESH_SECONDS = 15
DOUYIN_COOKIE_LOGIN_NAMES = {
    "sessionid",
    "sessionid_ss",
    "sid_guard",
    "sid_tt",
    "uid_tt",
    "uid_tt_ss",
    "passport_auth_status",
    "passport_auth_status_ss",
}


@dataclass
class DouyinCookieAutoState:
    lock: threading.Lock = field(default_factory=threading.Lock)
    page: Any = None
    browser_thread: Optional[threading.Thread] = None
    target_url: str = ""
    last_error: str = ""
    last_message: str = ""
    last_saved_at: float = 0.0
    last_cookie_fields: int = 0
    last_logged_in: bool = False

    def snapshot(self) -> dict:
        with self.lock:
            page = self.page
            browser_opened = _cookie_browser_alive(page)
            current_url = ""
            if browser_opened:
                try:
                    current_url = str(page.url or "")
                except Exception:
                    current_url = ""
            saved_at_ts = self.last_saved_at
            if saved_at_ts <= 0 and DOUYIN_COOKIE_FILE.exists():
                try:
                    saved_at_ts = DOUYIN_COOKIE_FILE.stat().st_mtime
                except OSError:
                    saved_at_ts = 0.0
            return {
                "available": _has_drissionpage(),
                "browser_opened": browser_opened,
                "target_url": self.target_url,
                "current_url": current_url,
                "auto_refresh_running": bool(self.browser_thread and self.browser_thread.is_alive()),
                "last_error": self.last_error,
                "last_message": self.last_message,
                "last_cookie_fields": self.last_cookie_fields,
                "logged_in": self.last_logged_in,
                "saved_at": saved_at_ts,
                "saved_at_text": (
                    datetime.fromtimestamp(saved_at_ts).strftime("%Y-%m-%d %H:%M:%S")
                    if saved_at_ts > 0
                    else ""
                ),
            }


DOUYIN_COOKIE_AUTO_STATE = DouyinCookieAutoState()


# ---------------------------------------------------------------------------
# Cookie 工具：支持字符串与 JSON 两种格式自适应
# ---------------------------------------------------------------------------

def _parse_cookie_input(raw: str) -> Tuple[str, int]:
    """把任意输入解析为标准 cookie 字符串，返回 (cookie_str, 字段数)。

    支持：
    - 浏览器复制的整段字符串："a=b; c=d; e=f"
    - JSON 对象：{"a": "b", "c": "d"} 或 [{"name":"a","value":"b"}, ...]（chrome cookie 导出格式）
    """
    text = (raw or "").strip()
    if not text:
        return "", 0

    # 尝试 JSON
    parsed = None
    try:
        parsed = json.loads(text)
    except Exception:
        parsed = None

    pairs: List[Tuple[str, str]] = []
    if isinstance(parsed, dict):
        for k, v in parsed.items():
            if k is None:
                continue
            pairs.append((str(k).strip(), str(v).strip()))
    elif isinstance(parsed, list):
        for item in parsed:
            if isinstance(item, dict):
                k = item.get("name") or item.get("key")
                v = item.get("value")
                if k:
                    pairs.append((str(k).strip(), str(v if v is not None else "").strip()))
    else:
        # 当作普通 cookie 字符串
        for chunk in text.split(";"):
            s = chunk.strip()
            if not s or "=" not in s:
                continue
            k, _, v = s.partition("=")
            pairs.append((k.strip(), v.strip()))

    # 去重保留最后一次出现
    seen: Dict[str, str] = {}
    for k, v in pairs:
        if k:
            seen[k] = v
    cookie_str = "; ".join(f"{k}={v}" for k, v in seen.items() if k)
    return cookie_str, len(seen)


def _save_douyin_cookie(raw: str) -> Tuple[str, int]:
    cookie_str, count = _parse_cookie_input(raw)
    if not cookie_str:
        raise ValueError("Cookie 解析后为空，请粘贴整段字符串或 JSON 对象")
    DOUYIN_COOKIE_FILE.parent.mkdir(parents=True, exist_ok=True)
    DOUYIN_COOKIE_FILE.write_text(cookie_str + "\n", encoding="utf-8")
    os.environ["DY_COOKIE"] = cookie_str
    return cookie_str, count


def _douyin_cookie_summary() -> dict:
    cookie_str = ""
    try:
        cookie_str = douyin_crawler.load_cookie()
    except Exception:
        cookie_str = os.getenv("DY_COOKIE") or os.getenv("DOUYIN_COOKIE") or ""
    saved_at = 0.0
    if DOUYIN_COOKIE_FILE.exists():
        try:
            saved_at = DOUYIN_COOKIE_FILE.stat().st_mtime
        except OSError:
            saved_at = 0.0
    parts = [p for p in (cookie_str or "").split(";") if p.strip() and "=" in p]
    keys = []
    for p in parts:
        k = p.split("=", 1)[0].strip()
        if k:
            keys.append(k)
    important = [
        k for k in (
            "UIFID", "ttwid", "sessionid", "sessionid_ss",
            "passport_csrf_token", "msToken", "s_v_web_id",
            "__ac_signature", "__ac_nonce",
        ) if k in keys
    ]
    return {
        "configured": bool(cookie_str.strip()),
        "length": len(cookie_str),
        "fields": len(keys),
        "important_fields_present": important,
        "has_uifid": "UIFID" in keys or "UIFID_TEMP" in keys,
        "has_ttwid": "ttwid" in keys,
        "has_session": "sessionid" in keys or "sessionid_ss" in keys,
        "has_verify_fp": "s_v_web_id" in keys or "verifyFp" in keys or "fp" in keys,
        "has_ac_signature": "__ac_signature" in keys and "__ac_nonce" in keys,
        "saved_at": saved_at,
        "saved_at_text": (
            datetime.fromtimestamp(saved_at).strftime("%Y-%m-%d %H:%M:%S")
            if saved_at > 0
            else ""
        ),
        "path": str(DOUYIN_COOKIE_FILE),
    }


def _has_drissionpage() -> bool:
    try:
        import DrissionPage  # noqa: F401

        return True
    except Exception:
        return False


def _cookie_user_agent() -> str:
    try:
        return str(douyin_crawler.room_api.USER_AGENT)
    except Exception:
        return (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/136.0.0.0 Safari/537.36"
        )


def _cookie_browser_alive(page: Any) -> bool:
    if page is None:
        return False
    try:
        states = getattr(page, "states", None)
        is_alive = getattr(states, "is_alive", None) if states is not None else None
        if isinstance(is_alive, bool):
            return is_alive
        _ = page.url
        return True
    except Exception:
        return False


def _cookie_store_from_header(cookie_header: str) -> Dict[str, str]:
    cookies: Dict[str, str] = {}
    for chunk in str(cookie_header or "").split(";"):
        item = chunk.strip()
        if not item or "=" not in item:
            continue
        k, v = item.split("=", 1)
        key = k.strip()
        if key:
            cookies[key] = v.strip()
    return cookies


def _cookie_has_login(cookies: Dict[str, str]) -> bool:
    return bool(set(cookies) & DOUYIN_COOKIE_LOGIN_NAMES)


def _cookie_normalize_items(cookie_items) -> List[Dict[str, str]]:
    if isinstance(cookie_items, dict):
        cookie_items = [
            {"name": name, "value": value, "domain": "douyin.com"}
            for name, value in cookie_items.items()
        ]
    normalized: List[Dict[str, str]] = []
    for cookie in cookie_items or []:
        if not isinstance(cookie, dict):
            continue
        name = str(cookie.get("name") or "").strip()
        value = cookie.get("value")
        domain = str(cookie.get("domain") or "")
        if not name or value is None:
            continue
        if domain and "douyin.com" not in domain and "bytedance.com" not in domain:
            continue
        normalized.append({
            "name": name,
            "value": str(value),
            "domain": domain,
        })
    return normalized


def _cookie_page_items(page) -> List[Dict[str, str]]:
    for kwargs in (
        {"as_dict": False, "all_domains": True, "all_info": True},
        {"all_domains": True},
        {},
    ):
        try:
            return _cookie_normalize_items(page.cookies(**kwargs))
        except TypeError:
            continue
    return []


def _cookie_header_from_page(page) -> str:
    items = _cookie_page_items(page)
    merged: Dict[str, str] = {}
    for item in items:
        merged[item["name"]] = item["value"]
    return "; ".join(f"{k}={v}" for k, v in merged.items())


def _cookie_target_url(keyword: str) -> str:
    key = str(keyword or "").strip()
    if key:
        return f"https://www.douyin.com/search/{urllib.parse.quote(key)}?type=live"
    return "https://www.douyin.com/?recommend=1"


def _build_cookie_browser():
    try:
        from DrissionPage import ChromiumOptions, ChromiumPage
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("缺少 DrissionPage 依赖，请先安装 requirements.txt 中新增的 DrissionPage") from exc

    options = ChromiumOptions()
    options.set_argument("--start-maximized")
    options.set_argument("--disable-gpu")
    options.set_argument("--no-sandbox")
    options.set_argument("--disable-dev-shm-usage")
    user_agent = _cookie_user_agent()
    if hasattr(options, "set_user_agent"):
        options.set_user_agent(user_agent)
    else:
        options.set_argument(f"--user-agent={user_agent}")
    return ChromiumPage(options)


def _apply_cookie_header_to_browser(page, cookie_header: str) -> None:
    cookie_dict = _cookie_store_from_header(cookie_header)
    if not cookie_dict:
        return
    formatted = [{"name": k, "value": v, "domain": ".douyin.com"} for k, v in cookie_dict.items()]
    try:
        page.set.cookies(formatted)
    except Exception:
        pass


def _save_cookie_from_browser(page, *, require_login: bool) -> Tuple[str, int, bool]:
    cookie_header = _cookie_header_from_page(page)
    cookie_dict = _cookie_store_from_header(cookie_header)
    logged_in = _cookie_has_login(cookie_dict)
    if require_login and not logged_in:
        raise ValueError("尚未检测到登录态 Cookie，请先扫码登录后再刷新获取")
    if not cookie_header:
        raise ValueError("未从自动化浏览器读取到任何 Cookie")
    if logged_in or require_login:
        _cookie_str, count = _save_douyin_cookie(cookie_header)
        return cookie_header, count, logged_in
    return cookie_header, len(cookie_dict), logged_in


def _cookie_refresh_worker() -> None:
    while True:
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            page = DOUYIN_COOKIE_AUTO_STATE.page
        if not _cookie_browser_alive(page):
            with DOUYIN_COOKIE_AUTO_STATE.lock:
                DOUYIN_COOKIE_AUTO_STATE.page = None
                DOUYIN_COOKIE_AUTO_STATE.last_message = "自动化浏览器已关闭"
            return
        try:
            _cookie_header, count, logged_in = _save_cookie_from_browser(page, require_login=False)
            with DOUYIN_COOKIE_AUTO_STATE.lock:
                DOUYIN_COOKIE_AUTO_STATE.last_cookie_fields = count
                DOUYIN_COOKIE_AUTO_STATE.last_logged_in = logged_in
                DOUYIN_COOKIE_AUTO_STATE.last_error = ""
                if logged_in:
                    DOUYIN_COOKIE_AUTO_STATE.last_saved_at = time.time()
                    DOUYIN_COOKIE_AUTO_STATE.last_message = f"已自动同步最新 Cookie（{count} 个字段）"
                else:
                    DOUYIN_COOKIE_AUTO_STATE.last_message = "自动化浏览器已打开，等待扫码登录"
        except Exception as exc:  # noqa: BLE001
            with DOUYIN_COOKIE_AUTO_STATE.lock:
                DOUYIN_COOKIE_AUTO_STATE.last_error = str(exc)
        time.sleep(DOUYIN_COOKIE_AUTO_REFRESH_SECONDS)


def _start_cookie_browser_session(keyword: str) -> dict:
    target_url = _cookie_target_url(keyword)
    with DOUYIN_COOKIE_AUTO_STATE.lock:
        page = DOUYIN_COOKIE_AUTO_STATE.page
        if page is not None and not _cookie_browser_alive(page):
            DOUYIN_COOKIE_AUTO_STATE.page = None
            page = None
        if page is None:
            page = _build_cookie_browser()
            existing_cookie = ""
            try:
                existing_cookie = douyin_crawler.load_cookie()
            except Exception:
                existing_cookie = os.getenv("DY_COOKIE") or ""
            if existing_cookie:
                _apply_cookie_header_to_browser(page, existing_cookie)
            DOUYIN_COOKIE_AUTO_STATE.page = page
            DOUYIN_COOKIE_AUTO_STATE.last_message = "自动化浏览器已启动"
        DOUYIN_COOKIE_AUTO_STATE.target_url = target_url
        DOUYIN_COOKIE_AUTO_STATE.last_error = ""

    page.get(target_url)
    try:
        page.set.window.max()
    except Exception:
        pass

    with DOUYIN_COOKIE_AUTO_STATE.lock:
        thread = DOUYIN_COOKIE_AUTO_STATE.browser_thread
        if not thread or not thread.is_alive():
            DOUYIN_COOKIE_AUTO_STATE.browser_thread = threading.Thread(
                target=_cookie_refresh_worker,
                daemon=True,
            )
            DOUYIN_COOKIE_AUTO_STATE.browser_thread.start()
        DOUYIN_COOKIE_AUTO_STATE.last_message = f"已打开抖音直播列表页，请扫码登录：{target_url}"
    return DOUYIN_COOKIE_AUTO_STATE.snapshot()


def _refresh_cookie_from_auto_browser() -> dict:
    with DOUYIN_COOKIE_AUTO_STATE.lock:
        page = DOUYIN_COOKIE_AUTO_STATE.page
    if not _cookie_browser_alive(page):
        raise RuntimeError("自动化浏览器未打开，请先点击“自动获取 Cookie”")
    _cookie_header, count, logged_in = _save_cookie_from_browser(page, require_login=True)
    with DOUYIN_COOKIE_AUTO_STATE.lock:
        DOUYIN_COOKIE_AUTO_STATE.last_cookie_fields = count
        DOUYIN_COOKIE_AUTO_STATE.last_logged_in = logged_in
        DOUYIN_COOKIE_AUTO_STATE.last_saved_at = time.time()
        DOUYIN_COOKIE_AUTO_STATE.last_error = ""
        DOUYIN_COOKIE_AUTO_STATE.last_message = f"已从自动化浏览器刷新并保存 Cookie（{count} 个字段）"
    return DOUYIN_COOKIE_AUTO_STATE.snapshot()


def _prepare_douyin_cookie_auto_page_for_monitor(keyword: str) -> bool:
    with DOUYIN_COOKIE_AUTO_STATE.lock:
        page = DOUYIN_COOKIE_AUTO_STATE.page
    if not _cookie_browser_alive(page):
        return False
    target_url = _cookie_target_url(keyword)
    try:
        page.get(target_url)
    except Exception as exc:  # noqa: BLE001
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            DOUYIN_COOKIE_AUTO_STATE.last_error = str(exc)
            DOUYIN_COOKIE_AUTO_STATE.last_message = f"复用 Cookie 页面跳转监控页失败: {exc}"
        return False
    try:
        page.run_js("window.focus && window.focus();")
    except Exception:
        pass
    try:
        _cookie_header, count, logged_in = _save_cookie_from_browser(page, require_login=False)
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            DOUYIN_COOKIE_AUTO_STATE.last_cookie_fields = count
            DOUYIN_COOKIE_AUTO_STATE.last_logged_in = logged_in
            DOUYIN_COOKIE_AUTO_STATE.target_url = target_url
            DOUYIN_COOKIE_AUTO_STATE.last_error = ""
            DOUYIN_COOKIE_AUTO_STATE.last_message = "已复用 Cookie 页面进入抖音监控页"
            if logged_in:
                DOUYIN_COOKIE_AUTO_STATE.last_saved_at = time.time()
    except Exception as exc:  # noqa: BLE001
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            DOUYIN_COOKIE_AUTO_STATE.target_url = target_url
            DOUYIN_COOKIE_AUTO_STATE.last_error = str(exc)
            DOUYIN_COOKIE_AUTO_STATE.last_message = f"已复用 Cookie 页面，但同步 Cookie 失败: {exc}"
    return True


def _cookie_page_has_verify(page: Any) -> bool:
    if not _cookie_browser_alive(page):
        return False
    checks = [
        "return (document.body && document.body.innerText ? document.body.innerText.slice(0, 5000) : '')",
        "return document.title || ''",
    ]
    text = ""
    for script in checks:
        try:
            text += str(page.run_js(script) or "")
        except Exception:
            continue
    return any(x in text for x in ("验证码", "请完成验证", "拖动滑块", "安全验证", "验证中心", "请先进行验证"))


def _fetch_douyin_rows_via_cookie_auto_page(keyword: str, *, max_pages: int = 1, wait_seconds: int = 25) -> List[Dict]:
    with DOUYIN_COOKIE_AUTO_STATE.lock:
        page = DOUYIN_COOKIE_AUTO_STATE.page
    if not _cookie_browser_alive(page):
        return []
    target_url = _cookie_target_url(keyword)
    current_url = ""
    try:
        current_url = str(page.url or "")
    except Exception:
        current_url = ""
    deadline = time.time() + max(5, wait_seconds)
    best: List[Dict] = []
    seen_keys: set = set()
    stable_rounds = 0
    last_count = -1
    target_count = (max_pages * 15) if max_pages and max_pages > 0 else 0
    state = {
        "response_count": 0,
        "has_more": True,
        "last_new_at": time.time(),
    }
    try:
        page.listen.start("aweme/v1/web/live/search/")
    except Exception:
        pass
    if current_url != target_url:
        page.get(target_url)
    else:
        try:
            page.refresh()
        except Exception:
            page.get(target_url)
    while time.time() < deadline:
        while True:
            try:
                packet = page.listen.wait(timeout=0.6)
            except Exception:
                packet = None
            if not packet:
                break
            try:
                payload = (packet.response.body if packet.response else None) or {}
            except Exception:
                payload = {}
            if not isinstance(payload, dict):
                continue
            state["response_count"] += 1
            nil_type = str((payload.get("search_nil_info") or {}).get("search_nil_type") or "")
            if nil_type == "verify_check":
                with DOUYIN_COOKIE_AUTO_STATE.lock:
                    DOUYIN_COOKIE_AUTO_STATE.last_message = "Cookie 页面处于验证码状态，请先完成验证"
                continue
            req_url = ""
            try:
                req_url = str(packet.request.url or "")
            except Exception:
                req_url = ""
            added = douyin_crawler.merge_live_rows(
                best,
                douyin_crawler.search_payload_to_rows(keyword, payload, req_url),
                seen_keys,
            )
            if added:
                state["last_new_at"] = time.time()
            if payload.get("has_more") in (0, False, "0"):
                state["has_more"] = False
        try:
            dom_rows = douyin_crawler._extract_search_rows_from_dom(page, keyword)
        except Exception:
            dom_rows = []
        dom_added = douyin_crawler.merge_live_rows(best, dom_rows, seen_keys)
        if dom_added:
            state["last_new_at"] = time.time()
        current_count = len(best)
        if current_count:
            if current_count == last_count:
                stable_rounds += 1
            else:
                stable_rounds = 0
                last_count = current_count
            if target_count and current_count >= target_count:
                break
            if state["response_count"] > 0 and not state["has_more"] and stable_rounds >= 2:
                break
            if not target_count and stable_rounds >= 4 and (time.time() - state["last_new_at"]) >= 4:
                break
            try:
                page.run_js("window.scrollBy(0, 2400);")
            except Exception:
                pass
        elif _cookie_page_has_verify(page):
            with DOUYIN_COOKIE_AUTO_STATE.lock:
                DOUYIN_COOKIE_AUTO_STATE.last_message = "Cookie 页面处于验证码状态，请先完成验证"
            time.sleep(1.2)
            continue
        else:
            try:
                page.run_js("window.scrollBy(0, 1800);")
            except Exception:
                pass
        time.sleep(1.2)
    try:
        page.listen.stop()
    except Exception:
        pass
    return best

app = Flask(__name__)

PLATFORMS = ("bilibili", "douyin")
PLATFORM_LABEL = {"bilibili": "B站", "douyin": "抖音"}
DEFAULT_KEYWORDS = ["示例关键词"]
DEFAULT_INTERVAL_MINUTES = 120
DEFAULT_INTERVAL_SECONDS = DEFAULT_INTERVAL_MINUTES * 60
DEFAULT_SUMMARY_TIME = "17:00"
DEFAULT_SUMMARY_RANGE_START = "00:00"
DEFAULT_SUMMARY_RANGE_END = "23:59"
DEFAULT_DAILY_RESET_TIME = DEFAULT_SUMMARY_TIME
DEFAULT_WECOM_WEBHOOK = ""
WECOM_CONFIG_FILE = BASE_DIR / "wecom_webhooks.json"
WECOM_CONFIG_LOCK = threading.Lock()
MONITOR_CONFIG_FILE = BASE_DIR / "monitor_configs.json"
MONITOR_CONFIG_LOCK = threading.Lock()
RUNTIME_STATE_FILE = BASE_DIR / "monitor_runtime_state.json"
RUNTIME_STATE_LOCK = threading.Lock()

EXPORT_HEADERS = [
    "抓取时间",
    "作者主页",
    "直播间",
    "主播",
    "标题",
    "开播时间",
    "看过人数",
    "实时点赞",
    "点赞来源",
    "高能/在线",
    "看过变化",
    "点赞变化",
    "分区",
]

DOUYIN_EXPORT_HEADERS = [
    "抓取时间",
    "搜索关键词",
    "作者主页",
    "直播间",
    "主播",
    "标题",
    "开播时间",
    "看过人数",
    "正在观看人数",
    "点赞数",
    "点赞来源",
    "看过变化",
    "点赞变化",
    "粉丝数",
    "粉丝团人数",
]


def _default_wecom_config() -> Dict[str, dict]:
    return {
        name: {
            "webhook_url": DEFAULT_WECOM_WEBHOOK,
            "updated_at": "",
            "auto_push": True,
        }
        for name in PLATFORMS
    }


def _normalize_wecom_config(data: Any) -> Dict[str, dict]:
    base = _default_wecom_config()
    if not isinstance(data, dict):
        return base
    for name in PLATFORMS:
        item = data.get(name) or {}
        if not isinstance(item, dict):
            item = {}
        webhook_url = str(item.get("webhook_url") or "").strip()
        if webhook_url:
            base[name]["webhook_url"] = webhook_url
        base[name]["updated_at"] = str(item.get("updated_at") or "").strip()
        base[name]["auto_push"] = bool(item.get("auto_push", True))
    return base


def _load_wecom_config() -> Dict[str, dict]:
    if not WECOM_CONFIG_FILE.exists():
        return _default_wecom_config()
    try:
        raw = json.loads(WECOM_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return _default_wecom_config()
    return _normalize_wecom_config(raw)


WECOM_CONFIG_CACHE = _load_wecom_config()


def _save_wecom_config() -> None:
    with WECOM_CONFIG_LOCK:
        WECOM_CONFIG_FILE.write_text(
            json.dumps(WECOM_CONFIG_CACHE, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def _mask_wecom_webhook(url: str) -> str:
    text = str(url or "").strip()
    if not text:
        return ""
    if len(text) <= 24:
        return text[:8] + "***"
    return text[:18] + "***" + text[-10:]


def _wecom_config_summary(platform: str) -> dict:
    with WECOM_CONFIG_LOCK:
        item = dict((WECOM_CONFIG_CACHE.get(platform) or {}))
    webhook_url = str(item.get("webhook_url") or "").strip()
    updated_at = str(item.get("updated_at") or "").strip()
    return {
        "platform": platform,
        "configured": bool(webhook_url),
        "webhook_url": webhook_url,
        "masked_webhook_url": _mask_wecom_webhook(webhook_url),
        "updated_at": updated_at,
        "auto_push": bool(item.get("auto_push", True)),
    }


def _set_wecom_config(
    platform: str,
    *,
    webhook_url: Optional[str] = None,
    auto_push: Optional[bool] = None,
) -> dict:
    with WECOM_CONFIG_LOCK:
        item = WECOM_CONFIG_CACHE.setdefault(platform, {"webhook_url": "", "updated_at": "", "auto_push": True})
        if webhook_url is not None:
            clean_url = str(webhook_url or "").strip()
            item["webhook_url"] = clean_url
            item["updated_at"] = now_str() if clean_url else ""
        if auto_push is not None:
            item["auto_push"] = bool(auto_push)
    _save_wecom_config()
    return _wecom_config_summary(platform)


def _default_saved_monitor_config(platform: str) -> dict:
    wecom = _wecom_config_summary(platform)
    return {
        "keywords": list(ADAPTERS[platform].default_keywords),
        "interval_minutes": DEFAULT_INTERVAL_MINUTES,
        "max_pages": 2,
        "enable_ws": True,
        "ws_workers": 16,
        "wecom_webhook_url": str(wecom.get("webhook_url") or ""),
        "wecom_auto_push": bool(wecom.get("auto_push", True)),
        "daily_reset_time": DEFAULT_SUMMARY_TIME,
        "continue_previous": False,
        "locked": False,
        "updated_at": "",
        "saved": False,
    }


def _normalize_saved_monitor_config(platform: str, item: Any) -> dict:
    base = _default_saved_monitor_config(platform)
    if not isinstance(item, dict):
        return base
    keywords = item.get("keywords")
    if isinstance(keywords, list):
        clean_keywords = [str(x).strip() for x in keywords if str(x).strip()]
        if clean_keywords:
            base["keywords"] = clean_keywords
    interval_minutes = max(1, min(24 * 60, parse_int(item.get("interval_minutes")) or DEFAULT_INTERVAL_MINUTES))
    max_pages = max(1, parse_int(item.get("max_pages")) or 2)
    ws_workers = max(2, min(64, parse_int(item.get("ws_workers")) or 16))
    daily_reset_time = normalize_daily_reset_time(item.get("daily_reset_time")) or DEFAULT_DAILY_RESET_TIME
    webhook_url = str(item.get("wecom_webhook_url") or "").strip() or str(base["wecom_webhook_url"])
    base.update({
        "interval_minutes": interval_minutes,
        "max_pages": max_pages,
        "enable_ws": bool(item.get("enable_ws", True)),
        "ws_workers": ws_workers,
        "wecom_webhook_url": webhook_url,
        "wecom_auto_push": bool(item.get("wecom_auto_push", base["wecom_auto_push"])),
        "daily_reset_time": daily_reset_time,
        "continue_previous": bool(item.get("continue_previous", False)),
        "locked": bool(item.get("locked", False)),
        "updated_at": str(item.get("updated_at") or "").strip(),
        "saved": bool(item.get("saved", False)),
    })
    return base


def _load_saved_monitor_configs() -> Dict[str, dict]:
    data = {name: _default_saved_monitor_config(name) for name in PLATFORMS}
    if not MONITOR_CONFIG_FILE.exists():
        return data
    try:
        raw = json.loads(MONITOR_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return data
    if not isinstance(raw, dict):
        return data
    for name in PLATFORMS:
        data[name] = _normalize_saved_monitor_config(name, raw.get(name))
    return data


SAVED_MONITOR_CONFIG_CACHE: Dict[str, dict] = {}


def _default_unified_monitor_config() -> dict:
    return {
        "platforms": list(PLATFORMS),
        "keywords": list(DEFAULT_KEYWORDS),
        "interval_minutes": DEFAULT_INTERVAL_MINUTES,
        "max_pages": 2,
        "enable_ws": True,
        "ws_workers": 16,
        "wecom_webhook_url": DEFAULT_WECOM_WEBHOOK,
        "wecom_auto_push": True,
        "summary_time": DEFAULT_SUMMARY_TIME,
        "summary_range_start": DEFAULT_SUMMARY_RANGE_START,
        "summary_range_end": DEFAULT_SUMMARY_RANGE_END,
        "blacklist_authors": [],
        "locked": False,
        "updated_at": "",
        "saved": False,
    }


def _normalize_platform_list(value: Any) -> List[str]:
    if isinstance(value, list):
        items = value
    else:
        items = [value]
    seen: set = set()
    result: List[str] = []
    for item in items:
        name = str(item or "").strip().lower()
        if name in PLATFORMS and name not in seen:
            seen.add(name)
            result.append(name)
    return result


def normalize_clock_time(value: Any) -> str:
    return normalize_daily_reset_time(value)


def normalize_author_home_url(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "://" not in text and not text.startswith("//"):
        text = f"https://{text}"
    parsed = urllib.parse.urlparse(text if "://" in text else f"https:{text}")
    scheme = "https"
    host = (parsed.netloc or "").strip().lower()
    path = re.sub(r"/+", "/", (parsed.path or "").strip())
    if not host and path:
        retry = urllib.parse.urlparse(f"https://{text.lstrip('/')}")
        host = (retry.netloc or "").strip().lower()
        path = re.sub(r"/+", "/", (retry.path or "").strip())
    host = host.removeprefix("www.")
    if not host:
        return ""
    path = path.rstrip("/")
    return f"{scheme}://{host}{path}"


def normalize_blacklist_authors(value: Any) -> List[str]:
    if isinstance(value, list):
        items = value
    else:
        text = str(value or "").replace("，", ",")
        for sep in ("\r\n", "\n", "\r", ",", ";", "；", "\t"):
            text = text.replace(sep, "\n")
        items = text.split("\n")
    seen: set = set()
    result: List[str] = []
    for item in items:
        normalized = normalize_author_home_url(item)
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def _load_unified_monitor_config() -> dict:
    base = _default_unified_monitor_config()
    if not MONITOR_CONFIG_FILE.exists():
        return base
    try:
        raw = json.loads(MONITOR_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return base
    source = raw.get("global") if isinstance(raw, dict) and isinstance(raw.get("global"), dict) else raw
    if not isinstance(source, dict):
        return base
    if "global" not in raw and any(name in source for name in PLATFORMS):
        legacy = source.get(PLATFORMS[0]) or next((source.get(name) for name in PLATFORMS if isinstance(source.get(name), dict)), {})
        if isinstance(legacy, dict):
            source = {
                "platforms": list(PLATFORMS),
                "keywords": legacy.get("keywords") or list(DEFAULT_KEYWORDS),
                "interval_minutes": legacy.get("interval_minutes") or DEFAULT_INTERVAL_MINUTES,
                "max_pages": legacy.get("max_pages") or 2,
                "enable_ws": legacy.get("enable_ws", True),
                "ws_workers": legacy.get("ws_workers") or 16,
                "wecom_webhook_url": legacy.get("wecom_webhook_url") or DEFAULT_WECOM_WEBHOOK,
                "wecom_auto_push": legacy.get("wecom_auto_push", True),
                "summary_time": legacy.get("daily_reset_time") or DEFAULT_SUMMARY_TIME,
                "summary_range_start": DEFAULT_SUMMARY_RANGE_START,
                "summary_range_end": DEFAULT_SUMMARY_RANGE_END,
                "blacklist_authors": legacy.get("blacklist_authors") or [],
                "updated_at": legacy.get("updated_at") or "",
                "saved": legacy.get("saved", False),
            }
    base["platforms"] = _normalize_platform_list(source.get("platforms")) or list(PLATFORMS)
    keywords = source.get("keywords")
    if isinstance(keywords, list):
        clean_keywords = [str(x).strip() for x in keywords if str(x).strip()]
        if clean_keywords:
            base["keywords"] = clean_keywords
    base["interval_minutes"] = max(1, min(24 * 60, parse_int(source.get("interval_minutes")) or DEFAULT_INTERVAL_MINUTES))
    base["max_pages"] = max(1, parse_int(source.get("max_pages")) or 2)
    base["enable_ws"] = bool(source.get("enable_ws", True))
    base["ws_workers"] = max(2, min(64, parse_int(source.get("ws_workers")) or 16))
    webhook_url = str(source.get("wecom_webhook_url") or "").strip()
    if webhook_url:
        base["wecom_webhook_url"] = webhook_url
    base["wecom_auto_push"] = bool(source.get("wecom_auto_push", True))
    base["summary_time"] = normalize_clock_time(source.get("summary_time")) or DEFAULT_SUMMARY_TIME
    base["summary_range_start"] = normalize_clock_time(source.get("summary_range_start")) or DEFAULT_SUMMARY_RANGE_START
    base["summary_range_end"] = normalize_clock_time(source.get("summary_range_end")) or DEFAULT_SUMMARY_RANGE_END
    base["blacklist_authors"] = normalize_blacklist_authors(source.get("blacklist_authors"))
    base["locked"] = bool(source.get("locked", False))
    base["updated_at"] = str(source.get("updated_at") or "").strip()
    base["saved"] = bool(source.get("saved", False))
    return base


UNIFIED_MONITOR_CONFIG_CACHE: Dict[str, Any] = {}


def _save_unified_monitor_config() -> None:
    with MONITOR_CONFIG_LOCK:
        payload = {"global": dict(UNIFIED_MONITOR_CONFIG_CACHE)}
        MONITOR_CONFIG_FILE.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def _unified_monitor_config_summary() -> dict:
    return {
        "platforms": list(UNIFIED_MONITOR_CONFIG_CACHE.get("platforms") or []),
        "keywords": list(UNIFIED_MONITOR_CONFIG_CACHE.get("keywords") or []),
        "interval_minutes": int(UNIFIED_MONITOR_CONFIG_CACHE.get("interval_minutes") or DEFAULT_INTERVAL_MINUTES),
        "max_pages": int(UNIFIED_MONITOR_CONFIG_CACHE.get("max_pages") or 2),
        "enable_ws": bool(UNIFIED_MONITOR_CONFIG_CACHE.get("enable_ws", True)),
        "ws_workers": int(UNIFIED_MONITOR_CONFIG_CACHE.get("ws_workers") or 16),
        "wecom_webhook_url": str(UNIFIED_MONITOR_CONFIG_CACHE.get("wecom_webhook_url") or ""),
        "wecom_auto_push": bool(UNIFIED_MONITOR_CONFIG_CACHE.get("wecom_auto_push", True)),
        "summary_time": str(UNIFIED_MONITOR_CONFIG_CACHE.get("summary_time") or DEFAULT_SUMMARY_TIME),
        "summary_range_start": str(UNIFIED_MONITOR_CONFIG_CACHE.get("summary_range_start") or DEFAULT_SUMMARY_RANGE_START),
        "summary_range_end": str(UNIFIED_MONITOR_CONFIG_CACHE.get("summary_range_end") or DEFAULT_SUMMARY_RANGE_END),
        "blacklist_authors": list(UNIFIED_MONITOR_CONFIG_CACHE.get("blacklist_authors") or []),
        "locked": bool(UNIFIED_MONITOR_CONFIG_CACHE.get("locked", False)),
        "updated_at": str(UNIFIED_MONITOR_CONFIG_CACHE.get("updated_at") or ""),
        "saved": bool(UNIFIED_MONITOR_CONFIG_CACHE.get("saved", False)),
    }


def _set_unified_monitor_config(data: dict) -> dict:
    cfg, selected_platforms = _build_monitor_config_from_payload(None, data)
    UNIFIED_MONITOR_CONFIG_CACHE.update({
        "platforms": selected_platforms,
        "keywords": list(cfg.keywords),
        "interval_minutes": int(cfg.interval_minutes),
        "max_pages": int(cfg.max_pages),
        "enable_ws": bool(cfg.enable_ws),
        "ws_workers": int(cfg.ws_workers),
        "wecom_webhook_url": str(cfg.wecom_webhook_url or ""),
        "wecom_auto_push": bool(cfg.wecom_auto_push),
        "summary_time": str(cfg.summary_time or DEFAULT_SUMMARY_TIME),
        "summary_range_start": str(cfg.summary_range_start or DEFAULT_SUMMARY_RANGE_START),
        "summary_range_end": str(cfg.summary_range_end or DEFAULT_SUMMARY_RANGE_END),
        "blacklist_authors": list(cfg.blacklist_author_links),
        "locked": True,
        "updated_at": now_str(),
        "saved": True,
    })
    _save_unified_monitor_config()
    if cfg.wecom_webhook_url:
        for platform in PLATFORMS:
            _set_wecom_config(platform, webhook_url=cfg.wecom_webhook_url, auto_push=cfg.wecom_auto_push)
    return _unified_monitor_config_summary()


def _set_unified_monitor_lock(locked: bool) -> dict:
    UNIFIED_MONITOR_CONFIG_CACHE["locked"] = bool(locked)
    if not UNIFIED_MONITOR_CONFIG_CACHE.get("saved"):
        UNIFIED_MONITOR_CONFIG_CACHE["updated_at"] = ""
    _save_unified_monitor_config()
    return _unified_monitor_config_summary()


def _save_saved_monitor_configs() -> None:
    with MONITOR_CONFIG_LOCK:
        MONITOR_CONFIG_FILE.write_text(
            json.dumps(SAVED_MONITOR_CONFIG_CACHE, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def _saved_monitor_config_summary(platform: str) -> dict:
    with MONITOR_CONFIG_LOCK:
        item = dict(SAVED_MONITOR_CONFIG_CACHE.get(platform) or _default_saved_monitor_config(platform))
    normalized = _normalize_saved_monitor_config(platform, item)
    return {
        "platform": platform,
        "locked": bool(normalized.get("locked", False)),
        "updated_at": str(normalized.get("updated_at") or ""),
        "saved": bool(normalized.get("saved", False)),
        "config": {
            "keywords": list(normalized.get("keywords") or []),
            "interval_minutes": int(normalized.get("interval_minutes") or DEFAULT_INTERVAL_MINUTES),
            "max_pages": int(normalized.get("max_pages") or 2),
            "enable_ws": bool(normalized.get("enable_ws", True)),
            "ws_workers": int(normalized.get("ws_workers") or 16),
            "wecom_webhook_url": str(normalized.get("wecom_webhook_url") or ""),
            "wecom_auto_push": bool(normalized.get("wecom_auto_push", True)),
            "daily_reset_time": str(normalized.get("daily_reset_time") or DEFAULT_DAILY_RESET_TIME),
            "continue_previous": bool(normalized.get("continue_previous", False)),
        },
    }


def _set_saved_monitor_config(platform: str, config_data: dict, *, locked: Optional[bool] = None) -> dict:
    normalized = _normalize_saved_monitor_config(platform, config_data)
    with MONITOR_CONFIG_LOCK:
        item = SAVED_MONITOR_CONFIG_CACHE.setdefault(platform, _default_saved_monitor_config(platform))
        item.update(normalized)
        item["saved"] = True
        item["updated_at"] = now_str()
        if locked is not None:
            item["locked"] = bool(locked)
    _save_saved_monitor_configs()
    return _saved_monitor_config_summary(platform)


def _set_saved_monitor_lock(platform: str, locked: bool) -> dict:
    with MONITOR_CONFIG_LOCK:
        item = SAVED_MONITOR_CONFIG_CACHE.setdefault(platform, _default_saved_monitor_config(platform))
        item["locked"] = bool(locked)
        if not item.get("saved"):
            item["updated_at"] = ""
    _save_saved_monitor_configs()
    return _saved_monitor_config_summary(platform)


def _extract_wecom_key(webhook_url: str) -> str:
    try:
        parsed = urllib.parse.urlparse(webhook_url)
        return urllib.parse.parse_qs(parsed.query).get("key", [""])[0]
    except Exception:
        return ""


def _send_wecom_text(webhook_url: str, content: str) -> Tuple[bool, str]:
    payload = {
        "msgtype": "text",
        "text": {"content": str(content or "")},
    }
    try:
        resp = requests.post(webhook_url, json=payload, timeout=15)
        data = resp.json()
    except Exception as exc:
        return False, f"企业微信文本发送异常: {exc}"
    if data.get("errcode") == 0:
        return True, "企业微信文本发送成功"
    return False, f"企业微信文本发送失败: {data}"


def _upload_wecom_file(webhook_url: str, file_path: Path) -> Tuple[Optional[str], str]:
    key = _extract_wecom_key(webhook_url)
    if not key:
        return None, "企业微信 webhook 缺少 key"
    upload_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file"
    try:
        with file_path.open("rb") as f:
            files = {"media": (file_path.name, f, "application/octet-stream")}
            resp = requests.post(upload_url, files=files, timeout=30)
        data = resp.json()
    except Exception as exc:
        return None, f"企业微信文件上传异常: {exc}"
    if data.get("errcode") == 0:
        return data.get("media_id"), "企业微信文件上传成功"
    return None, f"企业微信文件上传失败: {data}"


def _send_wecom_file(webhook_url: str, media_id: str) -> Tuple[bool, str]:
    payload = {
        "msgtype": "file",
        "file": {"media_id": media_id},
    }
    try:
        resp = requests.post(webhook_url, json=payload, timeout=15)
        data = resp.json()
    except Exception as exc:
        return False, f"企业微信文件发送异常: {exc}"
    if data.get("errcode") == 0:
        return True, "企业微信文件发送成功"
    return False, f"企业微信文件发送失败: {data}"


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_daily_reset_time(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})", text)
    if not m:
        return ""
    hour = int(m.group(1))
    minute = int(m.group(2))
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return ""
    return f"{hour:02d}:{minute:02d}"


def current_daily_reset_marker(reset_time: str, now: Optional[datetime] = None) -> str:
    normalized = normalize_daily_reset_time(reset_time)
    if not normalized:
        return ""
    dt_now = now or datetime.now()
    hour, minute = [int(x) for x in normalized.split(":", 1)]
    today_cutoff = dt_now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if dt_now < today_cutoff:
        today_cutoff -= timedelta(days=1)
    return today_cutoff.strftime("%Y-%m-%d %H:%M:%S")


def current_summary_marker(summary_time: str, now: Optional[datetime] = None) -> str:
    return current_daily_reset_marker(summary_time, now)


def _parse_dt(text: Any) -> Optional[datetime]:
    s = str(text or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def summary_period_bounds(
    marker_text: str,
    range_start: str,
    range_end: str,
) -> Tuple[Optional[datetime], Optional[datetime]]:
    marker_dt = _parse_dt(marker_text)
    start_text = normalize_clock_time(range_start)
    end_text = normalize_clock_time(range_end)
    if not marker_dt or not start_text or not end_text:
        return None, None
    start_h, start_m = [int(x) for x in start_text.split(":", 1)]
    end_h, end_m = [int(x) for x in end_text.split(":", 1)]
    end_dt = marker_dt.replace(hour=end_h, minute=end_m, second=59, microsecond=0)
    start_dt = marker_dt.replace(hour=start_h, minute=start_m, second=0, microsecond=0)
    # 起止时间相同也视为跨天整段，例如 17:00 -> 次日 17:00。
    if (end_h, end_m) <= (start_h, start_m):
        start_dt -= timedelta(days=1)
    if end_dt > marker_dt:
        end_dt = marker_dt
    return start_dt, end_dt


def room_key(row: dict) -> str:
    return f"{row.get('keyword', '')}::{row.get('room_id', '')}"


def parse_int(value) -> int:
    try:
        if value is None or value == "":
            return 0
        if isinstance(value, (int, float)):
            return int(float(value))
        s = str(value).strip().replace(",", "")
        if not s:
            return 0
        if "亿" in s:
            num = s.replace("亿", "").strip()
            return int(float(num) * 100000000)
        if "万" in s:
            num = s.replace("万", "").strip()
            return int(float(num) * 10000)
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if m:
            return int(float(m.group(0)))
        return 0
    except (TypeError, ValueError):
        return 0


def metric_value(row: dict, metric: str) -> int:
    if metric == "watched_count":
        return parse_int(row.get("watched_count"))
    if metric == "ws_like_count":
        return parse_int(row.get("ws_like_count"))
    if metric == "ws_online_rank_count":
        return parse_int(row.get("ws_online_rank_count"))
    if metric == "online":
        return parse_int(row.get("online"))
    return 0


def sort_online_value(row: dict) -> int:
    v = metric_value(row, "ws_online_rank_count")
    if v > 0:
        return v
    return metric_value(row, "online")


if not UNIFIED_MONITOR_CONFIG_CACHE:
    UNIFIED_MONITOR_CONFIG_CACHE.update(_load_unified_monitor_config())


# ---------------------------------------------------------------------------
# B 站平台抓取（沿用原 dashboard 的实现）
# ---------------------------------------------------------------------------

def format_live_start_time(room_info: dict) -> str:
    ts = room_info.get("live_start_time") or 0
    if isinstance(ts, (int, float)) and ts > 0:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    return ""


def get_room_info_data(session, room_id: int) -> dict:
    if room_id <= 0:
        return {}
    try:
        resp = session.get(
            bili_crawler.ROOM_INFO_URL,
            params={"room_id": room_id},
            timeout=10,
            headers={"Referer": f"https://live.bilibili.com/{room_id}"},
        )
        payload = resp.json() if resp.status_code == 200 else {}
        if payload.get("code") == 0:
            return payload.get("data") or {}
    except Exception:
        pass
    return {}


def enrich_row_from_room_api(session, row: dict) -> None:
    room_id = parse_int(row.get("room_id"))
    uid = parse_int(row.get("uid"))
    home = (row.get("anchor_homepage") or "").strip()
    if not home and uid:
        home = f"https://space.bilibili.com/{uid}"
        row["anchor_homepage"] = home
    row["author_home"] = home

    if room_id <= 0:
        row["live_start_time"] = str(row.get("live_time") or "").strip()
        return

    data = get_room_info_data(session, room_id)
    if not data:
        row["live_start_time"] = str(row.get("live_time") or "").strip()
        return

    room_info = data.get("room_info") or {}
    anchor_info = data.get("anchor_info") or {}
    anchor_base = anchor_info.get("base_info") or {}
    watched_show = data.get("watched_show") or {}

    lst = format_live_start_time(room_info)
    row["live_start_time"] = lst or str(row.get("live_time") or "").strip()

    parent = (room_info.get("parent_area_name") or "").strip()
    area = (room_info.get("area_name") or "").strip()
    merged_cat = "/".join(x for x in (parent, area) if x)
    if merged_cat:
        row["category"] = merged_cat

    uname = (anchor_base.get("uname") or "").strip()
    if uname:
        row["anchor_name"] = uname

    mid = room_info.get("uid") or anchor_base.get("mid")
    if mid:
        mid_int = int(mid)
        row["uid"] = mid_int
        row["author_home"] = f"https://space.bilibili.com/{mid_int}"
        row["anchor_homepage"] = row["author_home"]

    title = (room_info.get("title") or "").strip()
    if title:
        row["title"] = title

    if watched_show.get("num") is not None:
        row["watched_count"] = watched_show.get("num")

    oc = room_info.get("online")
    if oc is not None:
        row["online"] = oc


def bili_fetch_like_with_source(session, room_id: int) -> Tuple[Optional[int], str, str]:
    if room_id <= 0:
        return None, "未取到", "房间号无效"
    try:
        resp = session.get(
            bili_crawler.ROOM_INFO_URL,
            params={"room_id": room_id},
            timeout=10,
            headers={"Referer": f"https://live.bilibili.com/{room_id}"},
        )
        data = resp.json() if resp.status_code == 200 else {}
        if data.get("code") == 0:
            body = data.get("data") or {}
            like_info = body.get("like_info_v3") or body.get("like_info_v3_new") or {}
            val = like_info.get("click_count") or like_info.get("total_likes")
            if val is not None:
                return parse_int(val), "接口数据", "房间信息接口成功"
    except Exception:
        pass

    try:
        live_url = f"https://live.bilibili.com/{room_id}"
        resp = session.get(live_url, headers=bili_crawler.DEFAULT_HEADERS, timeout=15)
        if resp.status_code == 200:
            ssr = extract_ssr_data(resp.text)
            info_data = ssr.get("roomInfoRes", {}).get("data", {}) if isinstance(ssr, dict) else {}
            like_info = info_data.get("like_info_v3_new") or info_data.get("like_info_v3") or {}
            val = like_info.get("click_count") or like_info.get("total_likes")
            if val is not None:
                return parse_int(val), "页面数据", "直播页解析成功"
    except Exception:
        pass
    return None, "未取到", "接口和页面均未返回点赞"


def extract_ssr_data(html: str) -> dict:
    marker = "window.__NEPTUNE_IS_MY_WAIFU__="
    idx = html.find(marker)
    if idx < 0:
        return {}
    payload = html[idx + len(marker):]
    first = payload.find("{")
    if first < 0:
        return {}
    try:
        obj, _ = json.JSONDecoder().raw_decode(payload[first:])
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# 共用工具
# ---------------------------------------------------------------------------

def snapshot_sheet_title(fetched_at: str, node_no: int = 0) -> str:
    t = re.sub(r"[\[\]:*?/\\]", "_", (fetched_at or "").replace(":", "").replace(" ", "_"))
    base = f"节点_{t}".strip("_")
    if not t:
        base = f"节点_{node_no or 0:03d}"
    return base[:31]


def snapshot_cycle_from_title(title: str) -> int:
    m = re.match(r"^第(\d+)轮(?:_|$)", str(title or "").strip())
    return int(m.group(1)) if m else 0


def bili_row_to_export_cells(row: dict) -> List:
    author = (row.get("author_home") or row.get("anchor_homepage") or "").strip()
    return [
        row.get("fetched_at", ""),
        author,
        row.get("live_url", ""),
        row.get("anchor_name", ""),
        row.get("title", ""),
        row.get("live_start_time", "") or str(row.get("live_time") or "").strip(),
        metric_value(row, "watched_count"),
        metric_value(row, "ws_like_count"),
        row.get("like_source", ""),
        metric_value(row, "ws_online_rank_count"),
        parse_int(row.get("delta_watched")),
        parse_int(row.get("delta_like")),
        row.get("category", ""),
    ]


def douyin_row_to_export_cells(row: dict) -> List:
    author = (row.get("author_home") or row.get("anchor_homepage") or "").strip()
    return [
        row.get("fetched_at", ""),
        row.get("keyword", ""),
        author,
        row.get("live_url", ""),
        row.get("anchor_name", ""),
        row.get("title", ""),
        row.get("live_start_time", "") or str(row.get("live_time") or "").strip(),
        metric_value(row, "watched_count"),
        metric_value(row, "ws_online_rank_count"),
        metric_value(row, "ws_like_count"),
        row.get("like_source", ""),
        parse_int(row.get("delta_watched")),
        parse_int(row.get("delta_like")),
        parse_int(row.get("follower_count")),
        parse_int(row.get("fans_club_count")),
    ]


def bili_row_to_ui_public(row: dict) -> dict:
    author = (row.get("author_home") or row.get("anchor_homepage") or "").strip()
    return {
        "platform": "bilibili",
        "registry_key": room_key(row),
        "search_keyword": row.get("keyword", ""),
        "fetched_at": row.get("fetched_at", ""),
        "author_home": author,
        "live_url": row.get("live_url", ""),
        "anchor_name": row.get("anchor_name", ""),
        "title": row.get("title", ""),
        "live_start_time": row.get("live_start_time", "") or str(row.get("live_time") or "").strip(),
        "watched_count": metric_value(row, "watched_count"),
        "ws_like_count": metric_value(row, "ws_like_count"),
        "like_source": row.get("like_source", "") or "",
        "online_hot": metric_value(row, "ws_online_rank_count"),
        "delta_watched": parse_int(row.get("delta_watched")),
        "delta_like": parse_int(row.get("delta_like")),
        "category": row.get("category", ""),
        "extra": {},
    }


def douyin_row_to_ui_public(row: dict) -> dict:
    author = (row.get("author_home") or row.get("anchor_homepage") or "").strip()
    return {
        "platform": "douyin",
        "registry_key": room_key(row),
        "search_keyword": row.get("keyword", ""),
        "fetched_at": row.get("fetched_at", ""),
        "author_home": author,
        "live_url": row.get("live_url", ""),
        "anchor_name": row.get("anchor_name", ""),
        "title": row.get("title", ""),
        "live_start_time": row.get("live_start_time", "") or str(row.get("live_time") or "").strip(),
        "watched_count": metric_value(row, "watched_count"),
        "ws_like_count": metric_value(row, "ws_like_count"),
        "like_source": row.get("like_source", "") or "",
        "online_hot": metric_value(row, "ws_online_rank_count"),
        "delta_watched": parse_int(row.get("delta_watched")),
        "delta_like": parse_int(row.get("delta_like")),
        "category": "",
        "extra": {
            "follower_count": parse_int(row.get("follower_count")),
            "fans_club_count": parse_int(row.get("fans_club_count")),
            "web_rid": row.get("web_rid", ""),
        },
    }


def author_online_summary(unique_items: List[Tuple[str, dict]]) -> Tuple[Dict[str, dict], List[List]]:
    groups: Dict[str, dict] = {}

    for registry_key, u in unique_items:
        uid_int = parse_int(u.get("uid"))
        if uid_int > 0:
            gkey = f"uid:{uid_int}"
        else:
            home = (u.get("author_home") or "").strip()
            if home:
                gkey = f"home:{home}"
            else:
                name = (u.get("anchor_name") or "").strip()
                gkey = f"name:{name}" if name else f"room:{u.get('room_id', '')}"

        if gkey not in groups:
            groups[gkey] = {
                "display_uid": uid_int if uid_int > 0 else None,
                "anchor_name": "",
                "author_home": "",
                "rooms": {},
                "registry_keys": set(),
            }
        g = groups[gkey]
        rooms: Dict[str, int] = g["rooms"]
        g["registry_keys"].add(registry_key)
        rid = str(u.get("room_id", "")).strip()
        if not rid:
            continue
        avg_online = int(u.get("avg_online", u.get("max_online", 0)) or 0)
        rooms[rid] = max(rooms.get(rid, 0), avg_online)

        an = (u.get("anchor_name") or "").strip()
        if an and len(an) >= len(str(g.get("anchor_name") or "")):
            g["anchor_name"] = an
        ah = (u.get("author_home") or "").strip()
        if ah:
            g["author_home"] = ah

    registry_stats: Dict[str, dict] = {}
    out: List[List] = []
    for g in groups.values():
        rooms = g["rooms"]
        if not rooms:
            continue
        avg = sum(rooms.values()) / len(rooms)
        avg_i = int(round(avg))
        uid_cell = g.get("display_uid")
        uid_out: object = uid_cell if isinstance(uid_cell, int) and uid_cell > 0 else "—"
        row = [
            uid_out,
            g.get("anchor_name") or "",
            g.get("author_home") or "",
            len(rooms),
            avg_i,
        ]
        out.append(row)
        payload = {
            "author_room_count": len(rooms),
            "author_avg_online": avg_i,
        }
        for registry_key in g.get("registry_keys") or []:
            registry_stats[str(registry_key)] = dict(payload)

    out.sort(key=lambda row: (-parse_int(row[4]), str(row[0])))
    return registry_stats, out


def append_report_chart_sheets(
    wb: Workbook,
    snapshots: List[dict],
    unique_items: List[Tuple[str, dict]],
    author_lines: List[List],
) -> None:
    ws_data = wb.create_sheet("图表数据")
    ws_data.append(["节点序号", "节点时间", "直播间数", "看过人数合计", "点赞合计"])
    for snap in snapshots:
        rows = snap["rows"]
        ws_data.append([
            snap["cycle"],
            snap["fetched_at"],
            len(rows),
            sum(metric_value(r, "watched_count") for r in rows),
            sum(metric_value(r, "ws_like_count") for r in rows),
        ])
    n_snap = len(snapshots)
    snap_last = 1 + n_snap if n_snap else 1

    kw_title_row = snap_last + 2
    ws_data.cell(row=kw_title_row, column=1, value="按搜索关键词统计（累计条目数）")
    kw_header_row = kw_title_row + 1
    ws_data.cell(row=kw_header_row, column=1, value="搜索关键词")
    ws_data.cell(row=kw_header_row, column=2, value="直播间条目数")
    kw_counts: Dict[str, int] = {}
    for _k, u in unique_items:
        kw = str(u.get("search_keyword") or "—")
        kw_counts[kw] = kw_counts.get(kw, 0) + 1
    r_kw = kw_header_row + 1
    for kw, cnt in sorted(kw_counts.items(), key=lambda x: -x[1]):
        ws_data.cell(row=r_kw, column=1, value=kw)
        ws_data.cell(row=r_kw, column=2, value=cnt)
        r_kw += 1
    kw_last_row = r_kw - 1

    top_title_row = ws_data.max_row + 2
    ws_data.cell(row=top_title_row, column=1, value="直播间最高在线人数 Top15（主播）")
    top_header_row = top_title_row + 1
    ws_data.cell(row=top_header_row, column=1, value="主播")
    ws_data.cell(row=top_header_row, column=2, value="直播间最高在线人数")
    top_on = sorted(
        unique_items,
        key=lambda it: -int(it[1].get("max_online", 0) or 0),
    )[:15]
    r_top = top_header_row + 1
    for _key, u in top_on:
        name = str(u.get("anchor_name") or u.get("room_id") or "")[:32]
        ws_data.cell(row=r_top, column=1, value=name)
        ws_data.cell(row=r_top, column=2, value=int(u.get("max_online", 0) or 0))
        r_top += 1
    top_last_row = r_top - 1

    au_title_row = ws_data.max_row + 2
    ws_data.cell(row=au_title_row, column=1, value="按作者平均直播间在线人数 Top15")
    au_header_row = au_title_row + 1
    ws_data.cell(row=au_header_row, column=1, value="主播")
    ws_data.cell(row=au_header_row, column=2, value="作者维度平均在线人数")
    r_au = au_header_row + 1
    for line in author_lines[:15]:
        label = str(line[1] or line[0] or "")[:32]
        ws_data.cell(row=r_au, column=1, value=label)
        ws_data.cell(row=r_au, column=2, value=int(line[4]) if len(line) > 4 else 0)
        r_au += 1
    au_last_row = r_au - 1

    ws_vis = wb.create_sheet("汇报图表")
    ws_vis["A1"] = "以下为自动生成的汇报用图表，数据明细见「图表数据」表；可整体复制到 PPT / Word。"

    if n_snap >= 1:
        cat_ref = Reference(ws_data, min_col=1, min_row=2, max_row=1 + n_snap)
        lc1 = LineChart()
        lc1.title = "各节点抓取到的直播间数量"
        lc1.y_axis.title = "房间数"
        lc1.x_axis.title = "节点序号"
        lc1.style = 2
        lc1.height = 9
        lc1.width = 16
        d1 = Reference(ws_data, min_col=3, min_row=1, max_row=1 + n_snap, max_col=3)
        lc1.add_data(d1, titles_from_data=True)
        lc1.set_categories(cat_ref)
        ws_vis.add_chart(lc1, "A3")

        lc2 = LineChart()
        lc2.title = "各节点「看过人数」合计"
        lc2.y_axis.title = "看过人数合计"
        lc2.x_axis.title = "节点序号"
        lc2.style = 5
        lc2.height = 9
        lc2.width = 16
        d2 = Reference(ws_data, min_col=4, min_row=1, max_row=1 + n_snap, max_col=4)
        lc2.add_data(d2, titles_from_data=True)
        lc2.set_categories(cat_ref)
        ws_vis.add_chart(lc2, "K3")

    kw_first = kw_header_row + 1
    if kw_last_row >= kw_first:
        bar_kw = BarChart()
        bar_kw.barDir = "bar"
        bar_kw.title = "各搜索关键词下的直播间条目数"
        bar_kw.y_axis.title = "条目数"
        bar_kw.x_axis.title = "关键词"
        bar_kw.height = 9
        bar_kw.width = 16
        dk = Reference(ws_data, min_col=2, min_row=kw_header_row, max_row=kw_last_row, max_col=2)
        ck = Reference(ws_data, min_col=1, min_row=kw_first, max_row=kw_last_row)
        bar_kw.add_data(dk, titles_from_data=True)
        bar_kw.set_categories(ck)
        ws_vis.add_chart(bar_kw, "A25")

    top_first = top_header_row + 1
    if top_last_row >= top_first:
        col_top = BarChart()
        col_top.title = "直播间最高在线人数 Top15 主播"
        col_top.y_axis.title = "直播间最高在线人数"
        col_top.x_axis.title = "主播"
        col_top.height = 9
        col_top.width = 16
        dt = Reference(ws_data, min_col=2, min_row=top_header_row, max_row=top_last_row, max_col=2)
        ct = Reference(ws_data, min_col=1, min_row=top_first, max_row=top_last_row)
        col_top.add_data(dt, titles_from_data=True)
        col_top.set_categories(ct)
        ws_vis.add_chart(col_top, "K25")

    au_first = au_header_row + 1
    if au_last_row >= au_first and author_lines:
        bar_au = BarChart()
        bar_au.barDir = "bar"
        bar_au.title = "按作者：平均直播间在线人数（Top15）"
        bar_au.y_axis.title = "作者维度平均在线人数"
        bar_au.x_axis.title = "主播"
        bar_au.height = 9
        bar_au.width = 16
        da = Reference(ws_data, min_col=2, min_row=au_header_row, max_row=au_last_row, max_col=2)
        ca = Reference(ws_data, min_col=1, min_row=au_first, max_row=au_last_row)
        bar_au.add_data(da, titles_from_data=True)
        bar_au.set_categories(ca)
        ws_vis.add_chart(bar_au, "A48")

    any_chart = (
        n_snap >= 1
        or kw_last_row >= kw_first
        or top_last_row >= top_first
        or (au_last_row >= au_first and bool(author_lines))
    )
    if not any_chart:
        ws_vis["A3"] = "（当前无节点/无汇总数据，未生成图表；请先完成监控抓取后再导出。）"


# ---------------------------------------------------------------------------
# 平台适配器
# ---------------------------------------------------------------------------

@dataclass
class PlatformAdapter:
    name: str
    label: str
    discover_and_enrich: Callable
    export_headers: List[str]
    row_to_export_cells: Callable[[dict], List]
    row_to_ui_public: Callable[[dict], dict]
    default_keywords: List[str] = field(default_factory=list)


def _bili_discover_and_enrich(
    service: "LiveMonitorService", config: "MonitorConfig", cycle_rows: Dict[str, dict], prev_map: Dict[str, dict]
) -> None:
    session = bili_crawler.build_session()
    future_map: Dict[Any, dict] = {}
    workers = max(4, min(64, config.ws_workers))

    with ThreadPoolExecutor(max_workers=workers) as pool:
        for keyword in config.keywords:
            page = 1
            total_pages: Optional[int] = None
            seen_room_ids = set()

            while True:
                if not service.is_running():
                    return
                html = bili_crawler.fetch_search_page(session, keyword=keyword, page=page)
                live_rooms, _current_page, parsed_total_pages = bili_crawler.parse_live_room_page(html)

                if total_pages is None:
                    total_pages = parsed_total_pages
                    if config.max_pages:
                        total_pages = min(total_pages, config.max_pages)

                if not live_rooms:
                    break

                for item in live_rooms:
                    room_id = item.get("roomid")
                    if not room_id or room_id in seen_room_ids:
                        continue
                    seen_room_ids.add(room_id)
                    row = bili_crawler.convert_live_room(keyword=keyword, page=page, item=item)
                    row["platform"] = "bilibili"
                    enrich_row_from_room_api(session, row)
                    key = room_key(row)
                    if service.is_excluded(key):
                        continue
                    if config.enable_ws:
                        future = pool.submit(
                            bili_crawler.fetch_live_realtime_info,
                            session,
                            row.get("room_id"),
                            8.0,
                            row.get("uid") or None,
                        )
                        future_map[future] = row
                    else:
                        service._upsert_row_incremental(
                            row=row,
                            prev_row=prev_map.get(key, {}),
                            cycle_rows=cycle_rows,
                        )

                if total_pages is None or page >= total_pages:
                    break
                page += 1
                time.sleep(0.1)

        for future in as_completed(future_map):
            if not service.is_running():
                return
            row = future_map[future]
            try:
                info = future.result()
            except Exception:
                info = {}
            row["ws_online_rank_count"] = info.get("ws_online_rank_count")
            row["ws_like_count"] = info.get("ws_like_count")
            row["like_source"] = "实时消息"
            row["diag_realtime"] = (
                "成功" if row.get("ws_online_rank_count") not in (None, "") else "未返回"
            )
            row["diag_like"] = "成功" if row.get("ws_like_count") not in (None, "") else "未返回"
            row["diag_detail"] = "实时消息链路"
            if row.get("ws_like_count") in (None, ""):
                fallback_like, like_source, like_msg = bili_fetch_like_with_source(
                    session, parse_int(row.get("room_id"))
                )
                if fallback_like is not None:
                    row["ws_like_count"] = fallback_like
                    row["like_source"] = like_source
                    row["diag_like"] = "成功"
                    row["diag_detail"] = like_msg
            if row.get("ws_like_count") in (None, ""):
                row["ws_like_count"] = 0
                row["like_source"] = "未取到"
                row["diag_like"] = "失败"
                row["diag_detail"] = "点赞链路未返回"
            key = room_key(row)
            if service.is_excluded(key):
                continue
            service._upsert_row_incremental(
                row=row,
                prev_row=prev_map.get(key, {}),
                cycle_rows=cycle_rows,
            )


def _douyin_discover_and_enrich(
    service: "LiveMonitorService", config: "MonitorConfig", cycle_rows: Dict[str, dict], prev_map: Dict[str, dict]
) -> None:
    print(f"[抖音监听] 本轮开始，关键词={config.keywords} max_pages={config.max_pages} enable_ws={config.enable_ws}", flush=True)
    cookie = douyin_crawler.load_cookie()
    douyin_crawler.apply_user_agent_from_cookie(cookie)
    if config.keywords and not prev_map:
        first_keyword = next((str(k).strip() for k in config.keywords if str(k).strip()), "")
        if first_keyword:
            try:
                if _prepare_douyin_cookie_auto_page_for_monitor(first_keyword):
                    print(f"[抖音监听] 启动时复用 Cookie 页面: {first_keyword}", flush=True)
                else:
                    print(f"[抖音监听] 启动时预热独立浏览器页: {first_keyword}", flush=True)
                    douyin_crawler.prepare_search_browser_for_monitor(
                        first_keyword,
                        cookie=cookie,
                    )
            except Exception as exc:  # noqa: BLE001
                print(f"[抖音监听] 浏览器预热失败: {exc}", flush=True)
    sign_ctx = douyin_crawler.build_sign_ctx()
    session = douyin_crawler.build_session()

    future_map: Dict[Any, dict] = {}
    workers = max(2, min(16, config.ws_workers))
    total_found = 0
    empty_debugs: List[dict] = []

    with ThreadPoolExecutor(max_workers=workers) as pool:
        for keyword in config.keywords:
            if not service.is_running():
                return
            print(f"[抖音监听] 开始处理关键词: {keyword}", flush=True)
            rooms = []
            search_debug = None
            try:
                rooms, search_debug = douyin_crawler.fetch_search_all(
                    session,
                    keyword=keyword,
                    cookie=cookie,
                    sign_ctx=sign_ctx,
                    page_size=15,
                    max_pages=max(1, config.max_pages or 0),
                    return_debug=True,
                )
            except douyin_crawler.room_api.RiskControlError as exc:
                print(f"[抖音监听] HTTP 搜索触发风控/验证: {keyword} -> {exc}", flush=True)
                if "verify_check" not in str(exc):
                    service._set_last_error(f"抖音风控: {exc}")
                    continue
            except Exception as exc:  # noqa: BLE001
                print(f"[抖音监听] HTTP 搜索失败: {keyword} -> {exc}", flush=True)
                service._set_last_error(f"抖音搜索失败: {exc}")
                continue

            if not rooms:
                browser_rows = []
                if _prepare_douyin_cookie_auto_page_for_monitor(keyword):
                    print(f"[抖音监听] 进入同页 Cookie 浏览器兜底: {keyword}", flush=True)
                    try:
                        browser_rows = _fetch_douyin_rows_via_cookie_auto_page(
                            keyword=keyword,
                            max_pages=max(1, config.max_pages or 0),
                        )
                    except Exception as exc:  # noqa: BLE001
                        browser_rows = []
                        print(f"[抖音监听] 同页 Cookie 浏览器兜底失败: {keyword} -> {exc}", flush=True)
                if not browser_rows:
                    print(f"[抖音监听] 进入独立浏览器兜底: {keyword}", flush=True)
                    try:
                        browser_rows = douyin_crawler.fetch_search_rows_via_browser(
                            keyword=keyword,
                            cookie=cookie,
                            max_pages=max(1, config.max_pages or 0),
                        )
                    except Exception as exc:  # noqa: BLE001
                        browser_rows = []
                        print(f"[抖音监听] 浏览器兜底失败: {keyword} -> {exc}", flush=True)
                        service._set_last_error(f"抖音浏览器搜索失败: {exc}")
                if browser_rows:
                    print(f"[抖音监听] 浏览器兜底成功: {keyword} -> {len(browser_rows)} 个直播间", flush=True)
                    rooms = browser_rows

            if not rooms:
                print(f"[抖音监听] 关键词无结果: {keyword}", flush=True)
                empty_debugs.append(search_debug or {"keyword": keyword, "requests": []})
                continue
            total_found += len(rooms)
            print(f"[抖音监听] 关键词拿到直播间: {keyword} -> {len(rooms)} 个", flush=True)

            for raw in rooms:
                if isinstance(raw, dict) and raw.get("web_rid") and raw.get("platform") == "douyin":
                    row = dict(raw)
                else:
                    row = douyin_crawler.convert_live_room(keyword=keyword, page=1, item=raw)
                key = room_key(row)
                if service.is_excluded(key):
                    continue
                web_rid = row.get("web_rid") or ""
                if config.enable_ws and web_rid:
                    print(f"[抖音监听] 提交详情补全任务: web_rid={web_rid} anchor={row.get('anchor_name', '')}", flush=True)
                    future = pool.submit(
                        douyin_crawler.fetch_live_realtime_info,
                        session,
                        web_rid,
                        cookie,
                        sign_ctx,
                        row.get("sec_uid") or None,
                        fetch_profile=True,
                    )
                    future_map[future] = row
                else:
                    if not row.get("ws_like_count"):
                        row["ws_like_count"] = 0
                        row["like_source"] = "搜索数据"
                    else:
                        row["like_source"] = "搜索数据"
                    service._upsert_row_incremental(
                        row=row,
                        prev_row=prev_map.get(key, {}),
                        cycle_rows=cycle_rows,
                    )

    if total_found == 0 and empty_debugs:
        sample = next((item for item in empty_debugs if item), {})
        reqs = sample.get("requests") or []
        last_req = reqs[-1] if reqs else {}
        keyword = sample.get("keyword") or (config.keywords[0] if config.keywords else "")
        service._set_last_error(
            "抖音直播搜索列表为空："
            f"关键词「{keyword}」"
            f"HTTP {last_req.get('http_status', '—')} / "
            f"status_code={last_req.get('status_code', '—')} / "
            f"data_len={last_req.get('data_len', '—')} / "
            f"has_more={last_req.get('has_more', '—')}。"
            "当前 Cookie 已加载，更像是网页搜索接口被静默限制或签名参数失效，不是未登录。"
        )

    for future in as_completed(future_map):
        if not service.is_running():
            return
        row = future_map[future]
        try:
            info = future.result() or {}
            print(f"[抖音监听] 详情补全完成: web_rid={row.get('web_rid', '')} title={info.get('title', '')[:40]}", flush=True)
        except Exception:
            info = {}
            print(f"[抖音监听] 详情补全失败: web_rid={row.get('web_rid', '')}", flush=True)

        for k in (
            "anchor_name", "title", "live_start_time", "watched_count",
            "ws_online_rank_count", "ws_like_count", "online", "status",
            "status_str", "follower_count", "fans_club_count", "avatar",
            "author_home", "anchor_homepage", "uid", "room_id",
        ):
            v = info.get(k)
            if v not in (None, ""):
                row[k] = v
        if info.get("like_source"):
            row["like_source"] = info["like_source"]
        elif row.get("ws_like_count") in (None, "", 0):
            row["like_source"] = "未取到"
        else:
            row["like_source"] = row.get("like_source") or "搜索数据"
        if row.get("ws_like_count") in (None, ""):
            row["ws_like_count"] = 0

        key = room_key(row)
        if service.is_excluded(key):
            continue
        service._upsert_row_incremental(
            row=row,
            prev_row=prev_map.get(key, {}),
            cycle_rows=cycle_rows,
        )
    print(f"[抖音监听] 本轮结束，共写入 {len(cycle_rows)} 个直播间", flush=True)


ADAPTERS: Dict[str, PlatformAdapter] = {
    "bilibili": PlatformAdapter(
        name="bilibili",
        label="B站",
        discover_and_enrich=_bili_discover_and_enrich,
        export_headers=EXPORT_HEADERS,
        row_to_export_cells=bili_row_to_export_cells,
        row_to_ui_public=bili_row_to_ui_public,
        default_keywords=list(DEFAULT_KEYWORDS),
    ),
    "douyin": PlatformAdapter(
        name="douyin",
        label="抖音",
        discover_and_enrich=_douyin_discover_and_enrich,
        export_headers=DOUYIN_EXPORT_HEADERS,
        row_to_export_cells=douyin_row_to_export_cells,
        row_to_ui_public=douyin_row_to_ui_public,
        default_keywords=list(DEFAULT_KEYWORDS),
    ),
}


if not SAVED_MONITOR_CONFIG_CACHE:
    SAVED_MONITOR_CONFIG_CACHE.update(_load_saved_monitor_configs())


# ---------------------------------------------------------------------------
# 监控服务
# ---------------------------------------------------------------------------

@dataclass
class MonitorConfig:
    keywords: List[str] = field(default_factory=lambda: list(DEFAULT_KEYWORDS))
    interval_minutes: int = DEFAULT_INTERVAL_MINUTES
    interval_seconds: int = DEFAULT_INTERVAL_SECONDS
    max_pages: int = 2
    enable_ws: bool = True
    ws_workers: int = 16
    wecom_webhook_url: str = DEFAULT_WECOM_WEBHOOK
    wecom_auto_push: bool = True
    summary_time: str = DEFAULT_SUMMARY_TIME
    summary_range_start: str = DEFAULT_SUMMARY_RANGE_START
    summary_range_end: str = DEFAULT_SUMMARY_RANGE_END
    blacklist_author_links: List[str] = field(default_factory=list)


def _build_monitor_config_from_payload(platform: Optional[str], payload: dict) -> Tuple[MonitorConfig, List[str]]:
    default_keywords = list(ADAPTERS[platform].default_keywords) if platform in ADAPTERS else list(DEFAULT_KEYWORDS)
    keywords_raw = payload.get("keywords", ", ".join(default_keywords))
    if isinstance(keywords_raw, list):
        keywords = [str(k).strip() for k in keywords_raw if str(k).strip()]
    else:
        normalized_keywords = str(keywords_raw).replace("，", ",").replace("\n", ",")
        keywords = [k.strip() for k in normalized_keywords.split(",") if k.strip()]
    if not keywords:
        keywords = default_keywords

    interval_minutes = parse_int(payload.get("interval_minutes"))
    if interval_minutes <= 0:
        legacy_sec = parse_int(payload.get("interval_seconds"))
        interval_minutes = max(1, (legacy_sec + 59) // 60) if legacy_sec > 0 else DEFAULT_INTERVAL_MINUTES
    interval_minutes = max(1, min(24 * 60, interval_minutes))
    interval_seconds = interval_minutes * 60

    summary_time_raw = payload.get("summary_time", payload.get("daily_reset_time"))
    summary_time = normalize_clock_time(summary_time_raw)
    if str(summary_time_raw or "").strip() and not summary_time:
        raise ValueError("汇总时间格式不正确，请使用 HH:MM")
    if not summary_time:
        summary_time = DEFAULT_SUMMARY_TIME

    summary_range_start_raw = payload.get("summary_range_start")
    summary_range_end_raw = payload.get("summary_range_end")
    summary_range_start = normalize_clock_time(summary_range_start_raw)
    summary_range_end = normalize_clock_time(summary_range_end_raw)
    if str(summary_range_start_raw or "").strip() and not summary_range_start:
        raise ValueError("汇总起始时间格式不正确，请使用 HH:MM")
    if str(summary_range_end_raw or "").strip() and not summary_range_end:
        raise ValueError("汇总结束时间格式不正确，请使用 HH:MM")
    if not summary_range_start:
        summary_range_start = DEFAULT_SUMMARY_RANGE_START
    if not summary_range_end:
        summary_range_end = DEFAULT_SUMMARY_RANGE_END

    webhook_url = str(payload.get("wecom_webhook_url") or "").strip()
    auto_push_payload = payload.get("wecom_auto_push")
    auto_push = bool(auto_push_payload) if auto_push_payload is not None else bool(UNIFIED_MONITOR_CONFIG_CACHE.get("wecom_auto_push", True))
    if webhook_url and not webhook_url.startswith("https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key="):
        raise ValueError("企业微信 webhook 地址格式不正确")
    if webhook_url:
        for name in PLATFORMS:
            _set_wecom_config(name, webhook_url=webhook_url, auto_push=auto_push)
    else:
        saved_wecom = _unified_monitor_config_summary()
        webhook_url = str(saved_wecom.get("wecom_webhook_url") or "")
        auto_push = bool(saved_wecom.get("wecom_auto_push", True))

    platforms = _normalize_platform_list(payload.get("platforms"))
    if not platforms and platform in ADAPTERS:
        platforms = [str(platform)]
    if not platforms:
        platforms = list(UNIFIED_MONITOR_CONFIG_CACHE.get("platforms") or [])
    if not platforms:
        raise ValueError("请至少选择一个平台")

    cfg = MonitorConfig(
        keywords=keywords,
        interval_minutes=interval_minutes,
        interval_seconds=interval_seconds,
        max_pages=max(1, parse_int(payload.get("max_pages")) or 2),
        enable_ws=bool(payload.get("enable_ws", True)),
        ws_workers=max(2, min(64, parse_int(payload.get("ws_workers")) or 16)),
        wecom_webhook_url=webhook_url,
        wecom_auto_push=auto_push,
        summary_time=summary_time,
        summary_range_start=summary_range_start,
        summary_range_end=summary_range_end,
        blacklist_author_links=normalize_blacklist_authors(
            payload.get("blacklist_authors", payload.get("blacklist_author_links"))
        ),
    )
    return cfg, platforms


def _monitor_config_to_saved_dict(config: MonitorConfig) -> dict:
    return {
        "keywords": list(config.keywords),
        "interval_minutes": int(config.interval_minutes),
        "max_pages": int(config.max_pages),
        "enable_ws": bool(config.enable_ws),
        "ws_workers": int(config.ws_workers),
        "wecom_webhook_url": str(config.wecom_webhook_url or ""),
        "wecom_auto_push": bool(config.wecom_auto_push),
        "summary_time": str(config.summary_time or DEFAULT_SUMMARY_TIME),
        "summary_range_start": str(config.summary_range_start or DEFAULT_SUMMARY_RANGE_START),
        "summary_range_end": str(config.summary_range_end or DEFAULT_SUMMARY_RANGE_END),
        "blacklist_authors": list(config.blacklist_author_links),
    }


class LiveMonitorService:
    def __init__(self, adapter: PlatformAdapter):
        self.adapter = adapter
        self._lock = threading.Lock()
        self._running = False
        self._thread: Optional[threading.Thread] = None
        self._started_explicitly = False
        self._config = MonitorConfig()
        self._rows: List[dict] = []
        self._history: Dict[str, List[dict]] = {}
        self._excluded_keys: set = set()
        self._last_run_at = ""
        self._last_error = ""
        self._cycle_count = 0
        self._session_started_at = ""
        self._snapshots: List[dict] = []
        self._unique_registry: Dict[str, dict] = {}
        self._last_push_at = ""
        self._last_push_status = ""
        self._last_push_message = ""
        self._daily_reset_marker = ""
        self._previous_window_state: Optional[dict] = None

    def _has_runtime_data_locked(self) -> bool:
        return bool(self._snapshots) or bool(self._rows) or bool(self._unique_registry)

    def _snapshot_has_data(self, snapshot: Optional[dict]) -> bool:
        if not isinstance(snapshot, dict):
            return False
        return bool(snapshot.get("snapshots") or snapshot.get("rows") or snapshot.get("unique_registry"))

    def _snapshot_runtime_state_locked(self) -> dict:
        return {
            "rows": [dict(r) for r in self._rows],
            "history": {
                str(key): [dict(x) for x in values]
                for key, values in self._history.items()
            },
            "excluded_keys": list(self._excluded_keys),
            "last_run_at": self._last_run_at,
            "last_error": self._last_error,
            "cycle_count": self._cycle_count,
            "session_started_at": self._session_started_at,
            "snapshots": [dict(s, rows=[dict(x) for x in s["rows"]]) for s in self._snapshots],
            "unique_registry": {str(key): dict(value) for key, value in self._unique_registry.items()},
            "last_push_at": self._last_push_at,
            "last_push_status": self._last_push_status,
            "last_push_message": self._last_push_message,
        }

    def _clone_runtime_snapshot(self, snapshot: Optional[dict]) -> Optional[dict]:
        if not self._snapshot_has_data(snapshot):
            return None
        snap = snapshot or {}
        return {
            "rows": [dict(r) for r in snap.get("rows") or []],
            "history": {
                str(key): [dict(x) for x in values]
                for key, values in (snap.get("history") or {}).items()
            },
            "excluded_keys": list(snap.get("excluded_keys") or []),
            "last_run_at": str(snap.get("last_run_at") or ""),
            "last_error": str(snap.get("last_error") or ""),
            "cycle_count": int(snap.get("cycle_count") or 0),
            "session_started_at": str(snap.get("session_started_at") or ""),
            "snapshots": [dict(s, rows=[dict(x) for x in s["rows"]]) for s in (snap.get("snapshots") or [])],
            "unique_registry": {
                str(key): dict(value)
                for key, value in (snap.get("unique_registry") or {}).items()
            },
            "last_push_at": str(snap.get("last_push_at") or ""),
            "last_push_status": str(snap.get("last_push_status") or ""),
            "last_push_message": str(snap.get("last_push_message") or ""),
        }

    def _restore_runtime_state_locked(self, snapshot: Optional[dict]) -> bool:
        if not self._snapshot_has_data(snapshot):
            return False
        snap = snapshot or {}
        self._rows = [dict(r) for r in snap.get("rows") or []]
        self._history = {
            str(key): [dict(x) for x in values]
            for key, values in (snap.get("history") or {}).items()
        }
        self._excluded_keys = set(str(x) for x in (snap.get("excluded_keys") or []))
        self._last_run_at = str(snap.get("last_run_at") or "")
        self._last_error = str(snap.get("last_error") or "")
        self._cycle_count = int(snap.get("cycle_count") or 0)
        self._session_started_at = str(snap.get("session_started_at") or "")
        self._snapshots = [dict(s, rows=[dict(x) for x in s["rows"]]) for s in (snap.get("snapshots") or [])]
        self._unique_registry = {
            str(key): dict(value)
            for key, value in (snap.get("unique_registry") or {}).items()
        }
        self._last_push_at = str(snap.get("last_push_at") or "")
        self._last_push_status = str(snap.get("last_push_status") or "")
        self._last_push_message = str(snap.get("last_push_message") or "")
        return True

    def _reset_window_state_locked(self, session_started_at: str = ""):
        self._rows = []
        self._history = {}
        self._excluded_keys.clear()
        self._last_run_at = ""
        self._last_error = ""
        self._cycle_count = 0
        self._session_started_at = session_started_at
        self._snapshots = []
        self._unique_registry = {}
        self._last_push_at = ""
        self._last_push_status = ""
        self._last_push_message = ""

    def _clear_runtime_state_locked(self):
        self._reset_window_state_locked("")
        self._daily_reset_marker = ""
        self._started_explicitly = False

    @property
    def platform(self) -> str:
        return self.adapter.name

    def start(self, config: MonitorConfig, *, continue_previous: bool = False):
        thread_to_join: Optional[threading.Thread] = None
        with self._lock:
            self._config = config
            if self._running:
                self._running = False
                thread_to_join = self._thread
                self._thread = None

        if thread_to_join and thread_to_join.is_alive():
            thread_to_join.join(timeout=5)

        with self._lock:
            self._clear_runtime_state_locked()
            self._session_started_at = now_str()
            self._daily_reset_marker = current_summary_marker(config.summary_time)
            self._started_explicitly = True
            self._running = True
            self._thread = threading.Thread(target=self._loop, daemon=True)
            self._thread.start()
        _persist_runtime_states()

    def stop(self):
        thread_to_join: Optional[threading.Thread] = None
        with self._lock:
            self._running = False
            thread_to_join = self._thread
            self._thread = None
            self._started_explicitly = False

        if thread_to_join and thread_to_join.is_alive():
            thread_to_join.join(timeout=5)
        _persist_runtime_states()

    def reset(self):
        thread_to_join: Optional[threading.Thread] = None
        with self._lock:
            self._running = False
            thread_to_join = self._thread
            self._thread = None
            self._clear_runtime_state_locked()
            self._previous_window_state = None
        if thread_to_join and thread_to_join.is_alive():
            thread_to_join.join(timeout=5)
        _persist_runtime_states()

    def sanitize_for_api_state(self):
        with self._lock:
            should_reset = self._running and not self._started_explicitly
        if should_reset:
            self.reset()

    def is_running(self) -> bool:
        with self._lock:
            return self._running

    def is_excluded(self, key: str) -> bool:
        with self._lock:
            return key in self._excluded_keys

    def _set_last_error(self, err: str) -> None:
        with self._lock:
            self._last_error = err

    def _set_push_result(self, status: str, message: str) -> None:
        with self._lock:
            self._last_push_status = status
            self._last_push_message = message
            self._last_push_at = now_str() if status else ""

    def is_blacklisted_row(self, row: dict) -> bool:
        author_home = normalize_author_home_url(row.get("author_home") or row.get("anchor_homepage") or "")
        if not author_home:
            return False
        with self._lock:
            return author_home in set(self._config.blacklist_author_links or [])

    def _maybe_push_scheduled_summary(self, config: MonitorConfig) -> bool:
        marker = current_summary_marker(config.summary_time)
        if not marker:
            return False
        with self._lock:
            if not self._running:
                return False
            if not self._daily_reset_marker:
                self._daily_reset_marker = marker
                return False
            if marker == self._daily_reset_marker:
                return False
            self._daily_reset_marker = marker
            webhook_url = str(config.wecom_webhook_url or "").strip()
            auto_push = bool(config.wecom_auto_push)
        if not auto_push or not webhook_url:
            return False
        start_dt, end_dt = summary_period_bounds(
            marker,
            config.summary_range_start,
            config.summary_range_end,
        )
        if not start_dt or not end_dt:
            return False
        self.push_summary_to_wecom(
            webhook_url,
            start_at=start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            end_at=end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        )
        return True

    def _loop(self):
        while True:
            with self._lock:
                if not self._running:
                    break
                config = self._config
            try:
                self._run_once(config)
            except Exception as exc:  # noqa: BLE001
                with self._lock:
                    self._last_error = str(exc)
            sleep_seconds = max(60, config.interval_seconds)
            for _ in range(sleep_seconds):
                with self._lock:
                    if not self._running:
                        return
                self._maybe_push_scheduled_summary(config)
                time.sleep(1)

    def _run_once(self, config: MonitorConfig):
        with self._lock:
            self._last_error = ""
            prev_map = {room_key(r): r for r in self._rows}
        cycle_rows: Dict[str, dict] = {}
        self.adapter.discover_and_enrich(self, config, cycle_rows, prev_map)
        with self._lock:
            self._rows = sorted(cycle_rows.values(), key=sort_online_value, reverse=True)
            self._last_run_at = now_str()
            self._cycle_count += 1
            self._record_cycle_snapshot_locked()
            fetched_at = self._last_run_at
        _persist_runtime_states()
        self._push_current_node_to_wecom(config, fetched_at)
        self._maybe_push_scheduled_summary(config)

    def _build_wecom_summary_text(self, fetched_at: str, rows: List[dict], *, title: str = "当前节点推送") -> str:
        top_rows = sorted(rows, key=sort_online_value, reverse=True)
        total_rooms = len(rows)
        top_watched = sum(metric_value(r, "watched_count") for r in top_rows[:20])
        keywords = ", ".join(self._config.keywords) if self._config.keywords else "—"
        return (
            f"{self.adapter.label} {title}\n"
            f"节点时间: {fetched_at or now_str()}\n"
            f"关键词: {keywords}\n"
            f"当前直播间数: {total_rooms}\n"
            f"Top20 看过合计: {top_watched:,}"
        )

    def _push_current_node_to_wecom(self, config: MonitorConfig, fetched_at: str) -> None:
        if not config.wecom_auto_push:
            self._set_push_result("", "")
            return
        webhook_url = str(config.wecom_webhook_url or "").strip()
        if not webhook_url:
            self._set_push_result("", "")
            return
        self.push_current_to_wecom(webhook_url, fetched_at=fetched_at)

    def push_current_to_wecom(
        self,
        webhook_url: str,
        *,
        fetched_at: Optional[str] = None,
    ) -> Tuple[bool, str]:
        clean_url = str(webhook_url or "").strip()
        if not clean_url:
            return False, "企业微信 webhook 未配置"
        with self._lock:
            rows = [dict(r) for r in self._rows]
            pushed_at = fetched_at or self._last_run_at or now_str()
        summary_text = self._build_wecom_summary_text(pushed_at, rows, title="当前节点推送")
        if not rows:
            ok, msg = _send_wecom_text(
                clean_url,
                summary_text + "\n当前无直播间数据，未发送文件。",
            )
            self._set_push_result("success" if ok else "failed", msg)
            return ok, msg

        try:
            file_path = self.export_current_node_xlsx()
        except Exception as exc:
            msg = f"导出推送文件失败: {exc}"
            self._set_push_result("failed", msg)
            return False, msg

        media_id, upload_msg = _upload_wecom_file(clean_url, file_path)
        if not media_id:
            self._set_push_result("failed", upload_msg)
            return False, upload_msg

        file_ok, file_msg = _send_wecom_file(clean_url, media_id)
        text_ok, text_msg = _send_wecom_text(clean_url, summary_text)
        ok = file_ok and text_ok
        if ok:
            message = f"{file_msg}；{text_msg}"
            self._set_push_result("success", message)
            return True, message
        parts = [upload_msg]
        parts.append(file_msg)
        parts.append(text_msg)
        message = "；".join(p for p in parts if p)
        self._set_push_result("failed", message)
        return False, message

    def _snapshots_in_range(self, start_at: str, end_at: str) -> List[dict]:
        start_dt = _parse_dt(start_at)
        end_dt = _parse_dt(end_at)
        if not start_dt or not end_dt or start_dt > end_dt:
            return []
        with self._lock:
            snapshots = [dict(s, rows=[dict(x) for x in s["rows"]]) for s in self._snapshots]
        selected: List[dict] = []
        for snap in snapshots:
            fetched_dt = _parse_dt(snap.get("fetched_at"))
            if fetched_dt and start_dt <= fetched_dt <= end_dt:
                selected.append(snap)
        return selected

    def _build_unique_items_from_snapshots(self, snapshots: List[dict]) -> List[Tuple[str, dict]]:
        registry: Dict[str, dict] = {}
        for snap in snapshots:
            fetched_at = str(snap.get("fetched_at") or "")
            for row in snap.get("rows") or []:
                key = room_key(row)
                on_rank = metric_value(row, "ws_online_rank_count")
                on_simple = metric_value(row, "online")
                on_val = on_rank if on_rank > 0 else on_simple
                item = registry.get(key)
                if not item:
                    registry[key] = {
                        "search_keyword": row.get("keyword", ""),
                        "author_home": (row.get("author_home") or row.get("anchor_homepage") or "").strip(),
                        "room_id": row.get("room_id", ""),
                        "uid": row.get("uid", ""),
                        "anchor_name": row.get("anchor_name", ""),
                        "live_url": row.get("live_url", ""),
                        "live_start_time_last": row.get("live_start_time", "") or str(row.get("live_time") or "").strip(),
                        "category": row.get("category", ""),
                        "title_last": row.get("title", ""),
                        "platform": row.get("platform", self.platform),
                        "first_seen": fetched_at,
                        "last_seen": fetched_at,
                        "nodes_seen": 1,
                        "max_watched": metric_value(row, "watched_count"),
                        "max_like": metric_value(row, "ws_like_count"),
                        "max_online": on_val,
                        "online_total": on_val,
                        "avg_online": on_val,
                    }
                    continue
                item["last_seen"] = fetched_at
                item["nodes_seen"] = int(item.get("nodes_seen", 0)) + 1
                item["title_last"] = row.get("title", "") or item.get("title_last", "")
                item["anchor_name"] = row.get("anchor_name", "") or item.get("anchor_name", "")
                item["live_url"] = row.get("live_url", "") or item.get("live_url", "")
                item["author_home"] = (
                    (row.get("author_home") or row.get("anchor_homepage") or "").strip()
                    or item.get("author_home", "")
                )
                item["live_start_time_last"] = (
                    row.get("live_start_time", "")
                    or str(row.get("live_time") or "").strip()
                    or item.get("live_start_time_last", "")
                )
                item["category"] = row.get("category", "") or item.get("category", "")
                item["max_watched"] = max(int(item.get("max_watched", 0)), metric_value(row, "watched_count"))
                item["max_like"] = max(int(item.get("max_like", 0)), metric_value(row, "ws_like_count"))
                item["max_online"] = max(int(item.get("max_online", 0)), on_val)
                item["online_total"] = int(item.get("online_total", 0) or 0) + on_val
                nodes_seen = max(int(item.get("nodes_seen", 0) or 0), 1)
                item["avg_online"] = int(round(int(item.get("online_total", 0) or 0) / nodes_seen))
        return list(registry.items())

    def push_summary_to_wecom(self, webhook_url: str, *, start_at: str, end_at: str) -> Tuple[bool, str]:
        clean_url = str(webhook_url or "").strip()
        if not clean_url:
            return False, "企业微信 webhook 未配置"
        snapshots = self._snapshots_in_range(start_at, end_at)
        if not snapshots:
            msg = f"{self.adapter.label} 在所选节点范围内没有可汇总数据"
            self._set_push_result("failed", msg)
            return False, msg
        try:
            file_path = self.export_summary_range_xlsx(start_at=start_at, end_at=end_at)
        except Exception as exc:
            msg = f"汇总导出失败: {exc}"
            self._set_push_result("failed", msg)
            return False, msg
        latest_at = str(snapshots[-1].get("fetched_at") or end_at)
        row_count = sum(len(snap.get("rows") or []) for snap in snapshots)
        summary_text = self._build_wecom_summary_text(
            latest_at,
            [row for snap in snapshots for row in (snap.get("rows") or [])],
            title=f"节点汇总推送（{start_at} ~ {end_at}）",
        ) + f"\n节点数: {len(snapshots)}\n累计记录数: {row_count}"
        media_id, upload_msg = _upload_wecom_file(clean_url, file_path)
        if not media_id:
            self._set_push_result("failed", upload_msg)
            return False, upload_msg
        file_ok, file_msg = _send_wecom_file(clean_url, media_id)
        text_ok, text_msg = _send_wecom_text(clean_url, summary_text)
        ok = file_ok and text_ok
        message = f"{file_msg}；{text_msg}" if ok else "；".join(p for p in (upload_msg, file_msg, text_msg) if p)
        self._set_push_result("success" if ok else "failed", message)
        return ok, message

    def _record_cycle_snapshot_locked(self):
        rows_snapshot = [{**r} for r in self._rows]
        self._snapshots.append({
            "cycle": self._cycle_count,
            "fetched_at": self._last_run_at,
            "rows": rows_snapshot,
        })
        if len(self._snapshots) > 2000:
            del self._snapshots[:-2000]
        for row in self._rows:
            self._merge_unique_streamer_locked(row)

    def _merge_unique_streamer_locked(self, row: dict):
        key = room_key(row)
        t = self._last_run_at
        on_rank = metric_value(row, "ws_online_rank_count")
        on_simple = metric_value(row, "online")
        on_val = on_rank if on_rank > 0 else on_simple
        if key not in self._unique_registry:
            self._unique_registry[key] = {
                "search_keyword": row.get("keyword", ""),
                "author_home": (row.get("author_home") or row.get("anchor_homepage") or "").strip(),
                "room_id": row.get("room_id", ""),
                "uid": row.get("uid", ""),
                "anchor_name": row.get("anchor_name", ""),
                "live_url": row.get("live_url", ""),
                "live_start_time_last": row.get("live_start_time", "")
                or str(row.get("live_time") or "").strip(),
                "category": row.get("category", ""),
                "title_last": row.get("title", ""),
                "platform": row.get("platform", self.platform),
                "first_seen": t,
                "last_seen": t,
                "nodes_seen": 1,
                "max_watched": metric_value(row, "watched_count"),
                "max_like": metric_value(row, "ws_like_count"),
                "max_online": on_val,
                "online_total": on_val,
                "avg_online": on_val,
            }
            return
        u = self._unique_registry[key]
        u["last_seen"] = t
        u["nodes_seen"] = int(u.get("nodes_seen", 0)) + 1
        u["title_last"] = row.get("title", "") or u.get("title_last", "")
        u["anchor_name"] = row.get("anchor_name", "") or u.get("anchor_name", "")
        u["live_url"] = row.get("live_url", "") or u.get("live_url", "")
        u["author_home"] = (
            (row.get("author_home") or row.get("anchor_homepage") or "").strip()
            or u.get("author_home", "")
        )
        u["live_start_time_last"] = (
            row.get("live_start_time", "")
            or str(row.get("live_time") or "").strip()
            or u.get("live_start_time_last", "")
        )
        u["category"] = row.get("category", "") or u.get("category", "")
        u["max_watched"] = max(int(u.get("max_watched", 0)), metric_value(row, "watched_count"))
        u["max_like"] = max(int(u.get("max_like", 0)), metric_value(row, "ws_like_count"))
        u["max_online"] = max(int(u.get("max_online", 0)), on_val)
        u["online_total"] = int(u.get("online_total", 0) or 0) + on_val
        nodes_seen = max(int(u.get("nodes_seen", 0) or 0), 1)
        u["avg_online"] = int(round(int(u.get("online_total", 0) or 0) / nodes_seen))

    def _upsert_row_incremental(self, row: dict, prev_row: dict, cycle_rows: Dict[str, dict]):
        if self.is_blacklisted_row(row):
            return
        key = room_key(row)
        fetched_at = now_str()
        row["fetched_at"] = fetched_at
        row["platform"] = self.platform
        row["delta_watched"] = metric_value(row, "watched_count") - metric_value(prev_row, "watched_count")
        row["delta_like"] = metric_value(row, "ws_like_count") - metric_value(prev_row, "ws_like_count")

        with self._lock:
            hist = self._history.setdefault(key, [])
            hist.append({
                "time": fetched_at,
                "watched_count": metric_value(row, "watched_count"),
                "ws_like_count": metric_value(row, "ws_like_count"),
            })
            if len(hist) > 360:
                del hist[:-360]

            cycle_rows[key] = row
            self._rows = sorted(cycle_rows.values(), key=sort_online_value, reverse=True)
            self._last_run_at = fetched_at

    def exclude_rooms(self, keys: List[str]) -> dict:
        with self._lock:
            for key in keys:
                self._excluded_keys.add(key)
            self._rows = [r for r in self._rows if room_key(r) not in self._excluded_keys]
            return {"excluded_total": len(self._excluded_keys), "current_rooms": len(self._rows)}

    def clear_excluded(self) -> dict:
        with self._lock:
            self._excluded_keys.clear()
            return {"excluded_total": 0}

    def has_data(self) -> bool:
        with self._lock:
            return bool(self._snapshots) or bool(self._rows) or bool(self._unique_registry)

    def build_workbook(self, snapshots_override: Optional[List[dict]] = None, *, summary_title: str = "汇总") -> Workbook:
        with self._lock:
            snapshots = (
                [dict(s, rows=[dict(x) for x in s["rows"]]) for s in snapshots_override]
                if snapshots_override is not None
                else [dict(s, rows=[dict(x) for x in s["rows"]]) for s in self._snapshots]
            )
            session_started = self._session_started_at
            keywords = list(self._config.keywords)
            node_count = len(snapshots) if snapshots_override is not None else self._cycle_count
            exported_at = now_str()
            summary_time = self._config.summary_time
            summary_range_start = self._config.summary_range_start
            summary_range_end = self._config.summary_range_end
            blacklist_total = len(self._config.blacklist_author_links or [])
        unique_items = self._build_unique_items_from_snapshots(snapshots)

        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        used_titles = {summary_title}
        export_headers = self.adapter.export_headers
        cells_fn = self.adapter.row_to_export_cells

        for snap in snapshots:
            title = snapshot_sheet_title(snap["fetched_at"], int(snap.get("cycle") or 0))
            orig = title
            n = 2
            while title in used_titles:
                suf = f"_{n}"
                title = (orig[: 31 - len(suf)] + suf) if len(orig) + len(suf) > 31 else orig + suf
                n += 1
            used_titles.add(title)
            ws = wb.create_sheet(title=title)
            ws.append(list(export_headers))
            for row in snap["rows"]:
                ws.append(cells_fn(row))

        if not snapshots:
            ws0 = wb.create_sheet(title="暂无节点数据")
            ws0.append(["说明", "请先启动监控并完成至少一个节点抓取后再导出。"])

        ws_sum = wb.create_sheet(title=summary_title)
        ws_sum.append(["统计项", "内容"])
        ws_sum.append(["平台", self.adapter.label])
        ws_sum.append(["监控开始时间", session_started or "—"])
        ws_sum.append(["导出时间", exported_at])
        ws_sum.append(["搜索关键词", ", ".join(keywords) if keywords else "—"])
        ws_sum.append(["汇总时间", summary_time or "未设置"])
        ws_sum.append(["汇总时间段", f"{summary_range_start} ~ {summary_range_end}"])
        ws_sum.append(["黑名单作者数", blacklist_total])
        ws_sum.append(["已完成抓取节点数", node_count])
        ws_sum.append(["累计监听到不同直播间数", len(unique_items)])
        ws_sum.append([])

        author_stats_map, author_lines = author_online_summary(unique_items)
        detail_headers = [
            "平台",
            "搜索关键词",
            "作者主页",
            "房间号",
            "UID",
            "主播",
            "直播间链接",
            "末次开播时间",
            "分区",
            "首次出现",
            "最后出现",
            "出现节点数",
            "峰值看过",
            "峰值点赞",
            "直播间最高在线人数",
            "直播间平均在线人数",
            "作者维度直播间数",
            "作者维度平均在线人数",
            "末次标题",
        ]
        ws_sum.append(detail_headers)

        def sort_key(item: Tuple[str, dict]):
            _k, u = item
            return (-int(u.get("nodes_seen", 0)), str(u.get("last_seen", "")))

        for _key, u in sorted(unique_items, key=sort_key):
            author_stats = author_stats_map.get(str(_key), {})
            ws_sum.append([
                u.get("platform", self.platform),
                u.get("search_keyword", ""),
                u.get("author_home", ""),
                u.get("room_id", ""),
                u.get("uid", ""),
                u.get("anchor_name", ""),
                u.get("live_url", ""),
                u.get("live_start_time_last", ""),
                u.get("category", ""),
                u.get("first_seen", ""),
                u.get("last_seen", ""),
                u.get("nodes_seen", 0),
                u.get("max_watched", 0),
                u.get("max_like", 0),
                u.get("max_online", 0),
                u.get("avg_online", u.get("max_online", 0)),
                author_stats.get("author_room_count", 0),
                author_stats.get("author_avg_online", 0),
                u.get("title_last", ""),
            ])

        append_report_chart_sheets(wb, snapshots, unique_items, author_lines)
        return wb

    def export_single_xlsx(self) -> Path:
        wb = self.build_workbook()
        filename = (
            f"直播监控导出_{self.adapter.label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        file_path = EXPORT_DIR / filename
        wb.save(file_path)
        return file_path

    def export_current_node_xlsx(self) -> Path:
        with self._lock:
            if self._snapshots:
                snapshots = [dict(self._snapshots[-1], rows=[dict(x) for x in self._snapshots[-1]["rows"]])]
            else:
                snapshots = []
        wb = self.build_workbook(snapshots_override=snapshots, summary_title="当前节点")
        filename = (
            f"直播监控当前节点_{self.adapter.label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        file_path = EXPORT_DIR / filename
        wb.save(file_path)
        return file_path

    def export_summary_range_xlsx(self, *, start_at: str, end_at: str) -> Path:
        snapshots = self._snapshots_in_range(start_at, end_at)
        if not snapshots:
            raise ValueError("所选节点范围内没有可汇总的数据")
        wb = self.build_workbook(snapshots_override=snapshots, summary_title="节点汇总")
        filename = (
            f"直播监控节点汇总_{self.adapter.label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        file_path = EXPORT_DIR / filename
        wb.save(file_path)
        return file_path

    def state(self) -> dict:
        with self._lock:
            rows = list(self._rows)
            config = self._config
            running = self._running
            cycle_count = self._cycle_count
            last_run_at = self._last_run_at
            last_error = self._last_error
            session_started = self._session_started_at
            unique_n = len(self._unique_registry)
            snap_n = len(self._snapshots)
            excluded_count = len(self._excluded_keys)
            last_push_at = self._last_push_at
            last_push_status = self._last_push_status
            last_push_message = self._last_push_message
            available_nodes = [
                {
                    "value": str(s.get("fetched_at") or ""),
                    "label": str(s.get("fetched_at") or ""),
                }
                for s in reversed(self._snapshots)
                if str(s.get("fetched_at") or "")
            ]

        top_rows = sorted(rows, key=sort_online_value, reverse=True)
        summary = {
            "platform": self.platform,
            "label": self.adapter.label,
            "total_rooms": len(rows),
            "keywords": config.keywords,
            "running": running,
            "cycle_count": cycle_count,
            "node_count": cycle_count,
            "last_run_at": last_run_at,
            "last_error": last_error,
            "session_started_at": session_started,
            "unique_rooms_total": unique_n,
            "snapshots_stored": snap_n,
            "nodes_stored": snap_n,
            "interval_minutes": config.interval_minutes,
            "top_watched": sum(metric_value(r, "watched_count") for r in top_rows[:20]),
            "excluded_count": excluded_count,
            "blacklist_total": len(config.blacklist_author_links or []),
            "summary_time": config.summary_time,
            "summary_range_start": config.summary_range_start,
            "summary_range_end": config.summary_range_end,
            "wecom_configured": bool(str(config.wecom_webhook_url or "").strip()),
            "wecom_auto_push": bool(config.wecom_auto_push),
            "last_push_at": last_push_at,
            "last_push_status": last_push_status,
            "last_push_message": last_push_message,
        }

        return {
            "summary": summary,
            "config": config.__dict__,
            "current_rooms": [self.adapter.row_to_ui_public(r) for r in top_rows],
            "available_nodes": available_nodes,
        }

    def dump_persisted_state(self) -> dict:
        with self._lock:
            return {
                "platform": self.platform,
                "config": dict(self._config.__dict__),
                "running": bool(self._running),
                "started_explicitly": bool(self._started_explicitly),
                "daily_reset_marker": str(self._daily_reset_marker or ""),
                "current_window": self._snapshot_runtime_state_locked(),
                "previous_window": self._clone_runtime_snapshot(self._previous_window_state),
            }

    def load_persisted_state(self, data: Any) -> None:
        if not isinstance(data, dict):
            return
        config_raw = data.get("config") or {}
        if isinstance(config_raw, dict):
            interval_minutes = max(1, min(24 * 60, parse_int(config_raw.get("interval_minutes")) or DEFAULT_INTERVAL_MINUTES))
            self._config = MonitorConfig(
                keywords=[str(x).strip() for x in (config_raw.get("keywords") or []) if str(x).strip()] or list(self.adapter.default_keywords),
                interval_minutes=interval_minutes,
                interval_seconds=max(60, parse_int(config_raw.get("interval_seconds")) or interval_minutes * 60),
                max_pages=max(1, parse_int(config_raw.get("max_pages")) or 2),
                enable_ws=bool(config_raw.get("enable_ws", True)),
                ws_workers=max(2, min(64, parse_int(config_raw.get("ws_workers")) or 16)),
                wecom_webhook_url=str(config_raw.get("wecom_webhook_url") or DEFAULT_WECOM_WEBHOOK),
                wecom_auto_push=bool(config_raw.get("wecom_auto_push", True)),
                summary_time=normalize_clock_time(config_raw.get("summary_time", config_raw.get("daily_reset_time"))) or DEFAULT_SUMMARY_TIME,
                summary_range_start=normalize_clock_time(config_raw.get("summary_range_start")) or DEFAULT_SUMMARY_RANGE_START,
                summary_range_end=normalize_clock_time(config_raw.get("summary_range_end")) or DEFAULT_SUMMARY_RANGE_END,
                blacklist_author_links=normalize_blacklist_authors(config_raw.get("blacklist_authors")),
            )
        with self._lock:
            self._running = False
            self._thread = None
            self._started_explicitly = False
            self._clear_runtime_state_locked()
            self._daily_reset_marker = str(data.get("daily_reset_marker") or "")
            self._restore_runtime_state_locked(data.get("current_window"))
            self._previous_window_state = self._clone_runtime_snapshot(data.get("previous_window"))

    def restore_from_snapshots(self, config: "MonitorConfig", snapshots: List[dict], *, session_started_at: str = "") -> None:
        ordered = sorted(
            [dict(s) for s in snapshots if isinstance(s, dict)],
            key=lambda item: (int(item.get("cycle") or 0), str(item.get("fetched_at") or "")),
        )
        with self._lock:
            self._running = False
            self._thread = None
            self._started_explicitly = False
            self._config = config
            self._clear_runtime_state_locked()
            self._daily_reset_marker = current_summary_marker(config.summary_time)
            self._session_started_at = session_started_at or (str(ordered[0].get("fetched_at") or "") if ordered else now_str())
            self._previous_window_state = None
            for snap in ordered:
                cycle_no = max(1, int(snap.get("cycle") or (self._cycle_count + 1)))
                fetched_at = str(snap.get("fetched_at") or "")
                rows = [dict(r) for r in (snap.get("rows") or []) if isinstance(r, dict)]
                self._rows = sorted(rows, key=sort_online_value, reverse=True)
                self._last_run_at = fetched_at
                self._cycle_count = cycle_no
                self._snapshots.append({
                    "cycle": cycle_no,
                    "fetched_at": fetched_at,
                    "rows": [dict(r) for r in rows],
                })
                self._history = {}
                for row in rows:
                    key = room_key(row)
                    self._history.setdefault(key, []).append({
                        "time": fetched_at,
                        "watched_count": metric_value(row, "watched_count"),
                        "ws_like_count": metric_value(row, "ws_like_count"),
                    })
                    self._merge_unique_streamer_locked(row)


def _safe_text(value: Any) -> str:
    return str(value or "").strip()


def _resume_export_files(platform: str) -> List[Path]:
    prefix = f"直播监控导出_{ADAPTERS[platform].label}_"
    files = [p for p in EXPORT_DIR.glob("*.xlsx") if p.name.startswith(prefix)]
    return sorted(files, key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)


def _resume_summary_from_sheet(ws) -> Tuple[dict, List[dict]]:
    summary: dict = {}
    details: List[dict] = []
    mode = "summary"
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        first = _safe_text(values[0] if len(values) > 0 else "")
        second = _safe_text(values[1] if len(values) > 1 else "")
        if mode == "summary":
            if first == "平台" and second == "搜索关键词":
                mode = "detail"
                continue
            if first:
                summary[first] = second
            continue
        if not any(v not in (None, "") for v in values):
            continue
        details.append({
            "platform": first,
            "search_keyword": second,
            "author_home": _safe_text(values[2] if len(values) > 2 else ""),
            "room_id": _safe_text(values[3] if len(values) > 3 else ""),
            "uid": _safe_text(values[4] if len(values) > 4 else ""),
            "anchor_name": _safe_text(values[5] if len(values) > 5 else ""),
            "live_url": _safe_text(values[6] if len(values) > 6 else ""),
            "live_start_time_last": _safe_text(values[7] if len(values) > 7 else ""),
            "category": _safe_text(values[8] if len(values) > 8 else ""),
            "first_seen": _safe_text(values[9] if len(values) > 9 else ""),
            "last_seen": _safe_text(values[10] if len(values) > 10 else ""),
            "rounds_seen": parse_int(values[11] if len(values) > 11 else 0),
            "max_watched": parse_int(values[12] if len(values) > 12 else 0),
            "max_like": parse_int(values[13] if len(values) > 13 else 0),
            "max_online": parse_int(values[14] if len(values) > 14 else 0),
            "avg_online": parse_int(values[15] if len(values) > 15 else 0),
            "title_last": _safe_text(values[18] if len(values) > 18 else ""),
        })
    return summary, details


def _resume_detail_lookup(details: List[dict]) -> Dict[str, dict]:
    lookup: Dict[str, dict] = {}
    for item in details:
        live_url = _safe_text(item.get("live_url"))
        author_home = _safe_text(item.get("author_home"))
        anchor_name = _safe_text(item.get("anchor_name"))
        for key in (
            f"live_url::{live_url}",
            f"author::{author_home}::{anchor_name}",
            f"author::{author_home}",
        ):
            if key and key not in lookup:
                lookup[key] = item
    return lookup


def _resume_match_detail(lookup: Dict[str, dict], *, live_url: str, author_home: str, anchor_name: str) -> dict:
    for key in (
        f"live_url::{_safe_text(live_url)}",
        f"author::{_safe_text(author_home)}::{_safe_text(anchor_name)}",
        f"author::{_safe_text(author_home)}",
    ):
        if key in lookup:
            return lookup[key]
    return {}


def _extract_room_id_from_live_url(platform: str, live_url: str) -> str:
    url = _safe_text(live_url)
    if not url:
        return ""
    if platform == "bilibili":
        m = re.search(r"live\.bilibili\.com/(\d+)", url)
        return m.group(1) if m else ""
    m = re.search(r"live\.douyin\.com/([0-9]+)", url)
    return m.group(1) if m else ""


def _extract_uid_from_author_home(platform: str, author_home: str) -> str:
    url = _safe_text(author_home)
    if not url:
        return ""
    if platform == "bilibili":
        m = re.search(r"/(\d+)(?:[/?#]|$)", url)
        return m.group(1) if m else ""
    m = re.search(r"/user/([^/?#]+)", url)
    return m.group(1) if m else ""


def _parse_resume_snapshot_rows(platform: str, ws, detail_lookup: Dict[str, dict]) -> List[dict]:
    rows: List[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        values = list(raw)
        if not any(v not in (None, "") for v in values):
            continue
        if platform == "bilibili":
            author_home = _safe_text(values[1] if len(values) > 1 else "")
            live_url = _safe_text(values[2] if len(values) > 2 else "")
            anchor_name = _safe_text(values[3] if len(values) > 3 else "")
            meta = _resume_match_detail(detail_lookup, live_url=live_url, author_home=author_home, anchor_name=anchor_name)
            row = {
                "platform": platform,
                "fetched_at": _safe_text(values[0] if len(values) > 0 else ""),
                "author_home": author_home,
                "anchor_homepage": author_home,
                "live_url": live_url,
                "anchor_name": anchor_name,
                "title": _safe_text(values[4] if len(values) > 4 else ""),
                "live_start_time": _safe_text(values[5] if len(values) > 5 else ""),
                "watched_count": parse_int(values[6] if len(values) > 6 else 0),
                "ws_like_count": parse_int(values[7] if len(values) > 7 else 0),
                "like_source": _safe_text(values[8] if len(values) > 8 else ""),
                "ws_online_rank_count": parse_int(values[9] if len(values) > 9 else 0),
                "delta_watched": parse_int(values[10] if len(values) > 10 else 0),
                "delta_like": parse_int(values[11] if len(values) > 11 else 0),
                "category": _safe_text(values[12] if len(values) > 12 else "") or _safe_text(meta.get("category")),
                "keyword": _safe_text(meta.get("search_keyword")),
                "room_id": _safe_text(meta.get("room_id")) or _extract_room_id_from_live_url(platform, live_url),
                "uid": _safe_text(meta.get("uid")) or _extract_uid_from_author_home(platform, author_home),
            }
        else:
            author_home = _safe_text(values[2] if len(values) > 2 else "")
            live_url = _safe_text(values[3] if len(values) > 3 else "")
            anchor_name = _safe_text(values[4] if len(values) > 4 else "")
            meta = _resume_match_detail(detail_lookup, live_url=live_url, author_home=author_home, anchor_name=anchor_name)
            online_val = parse_int(values[8] if len(values) > 8 else 0)
            row = {
                "platform": platform,
                "fetched_at": _safe_text(values[0] if len(values) > 0 else ""),
                "keyword": _safe_text(values[1] if len(values) > 1 else "") or _safe_text(meta.get("search_keyword")),
                "author_home": author_home,
                "anchor_homepage": author_home,
                "live_url": live_url,
                "anchor_name": anchor_name,
                "title": _safe_text(values[5] if len(values) > 5 else ""),
                "live_start_time": _safe_text(values[6] if len(values) > 6 else ""),
                "watched_count": parse_int(values[7] if len(values) > 7 else 0),
                "online": online_val,
                "ws_online_rank_count": online_val,
                "ws_like_count": parse_int(values[9] if len(values) > 9 else 0),
                "like_source": _safe_text(values[10] if len(values) > 10 else ""),
                "delta_watched": parse_int(values[11] if len(values) > 11 else 0),
                "delta_like": parse_int(values[12] if len(values) > 12 else 0),
                "follower_count": parse_int(values[13] if len(values) > 13 else 0),
                "fans_club_count": parse_int(values[14] if len(values) > 14 else 0),
                "room_id": _safe_text(meta.get("room_id")) or _extract_room_id_from_live_url(platform, live_url),
                "uid": _safe_text(meta.get("uid")) or _extract_uid_from_author_home(platform, author_home),
            }
        rows.append(row)
    return rows


def _resume_export_options(platform: str) -> List[dict]:
    options: List[dict] = []
    for path in _resume_export_files(platform):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            nodes = []
            for title in wb.sheetnames:
                cycle = snapshot_cycle_from_title(title)
                if cycle <= 0:
                    continue
                fetched_at = ""
                try:
                    ws = wb[title]
                    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                        fetched_at = _safe_text(row[0] if row else "")
                        break
                except Exception:
                    fetched_at = ""
                nodes.append({
                    "sheet_title": title,
                    "cycle": cycle,
                    "fetched_at": fetched_at,
                    "label": f"第{cycle}轮 {fetched_at}".strip(),
                })
        finally:
            wb.close()
        if not nodes:
            continue
        options.append({
            "file_name": path.name,
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "nodes": sorted(nodes, key=lambda item: item["cycle"], reverse=True),
        })
    return options


def _build_resume_config_from_summary(platform: str, summary_map: dict) -> MonitorConfig:
    saved = _unified_monitor_config_summary()
    try:
        cfg, _ = _build_monitor_config_from_payload(platform, saved)
    except Exception:
        cfg = MonitorConfig(keywords=list(ADAPTERS[platform].default_keywords))
    keywords_text = _safe_text(summary_map.get("搜索关键词"))
    if keywords_text:
        cfg.keywords = [x.strip() for x in keywords_text.replace("，", ",").split(",") if x.strip()]
    summary_time = normalize_clock_time(summary_map.get("汇总时间") or summary_map.get("每日重置时间"))
    if summary_time:
        cfg.summary_time = summary_time
    return cfg


def _restore_service_from_export(platform: str, file_name: str, sheet_title: str) -> dict:
    safe_name = Path(file_name).name
    path = EXPORT_DIR / safe_name
    if not path.exists() or safe_name not in [p.name for p in _resume_export_files(platform)]:
        raise ValueError("所选导出文件不存在或与平台不匹配")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"导出文件读取失败: {exc}") from exc
    try:
        if sheet_title not in wb.sheetnames:
            raise ValueError("所选节点不存在")
        target_cycle = snapshot_cycle_from_title(sheet_title)
        if target_cycle <= 0:
            raise ValueError("所选节点不是有效轮次")
        if "总结" not in wb.sheetnames:
            raise ValueError("导出文件缺少总结页，无法恢复")
        summary_map, details = _resume_summary_from_sheet(wb["总结"])
        detail_lookup = _resume_detail_lookup(details)
        snapshots: List[dict] = []
        for title in wb.sheetnames:
            cycle = snapshot_cycle_from_title(title)
            if cycle <= 0 or cycle > target_cycle:
                continue
            rows = _parse_resume_snapshot_rows(platform, wb[title], detail_lookup)
            if not rows:
                continue
            fetched_at = str(rows[0].get("fetched_at") or "")
            snapshots.append({
                "cycle": cycle,
                "fetched_at": fetched_at,
                "rows": rows,
            })
        if not snapshots:
            raise ValueError("所选节点没有可恢复的直播数据")
    finally:
        wb.close()
    cfg = _build_resume_config_from_summary(platform, summary_map)
    SERVICES[platform].restore_from_snapshots(
        cfg,
        snapshots,
        session_started_at=_safe_text(summary_map.get("监控开始时间")),
    )
    _persist_runtime_states()
    return {
        "file_name": safe_name,
        "sheet_title": sheet_title,
        "cycle_count": max(int(s.get("cycle") or 0) for s in snapshots),
        "restored_snapshots": len(snapshots),
        "config": cfg.__dict__,
        "service": SERVICES[platform].state(),
    }


def _load_runtime_state_cache() -> Dict[str, dict]:
    if not RUNTIME_STATE_FILE.exists():
        return {}
    try:
        raw = json.loads(RUNTIME_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return raw if isinstance(raw, dict) else {}


def _save_runtime_state_cache(data: Dict[str, dict]) -> None:
    with RUNTIME_STATE_LOCK:
        RUNTIME_STATE_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def _persist_runtime_states() -> None:
    data = {name: service.dump_persisted_state() for name, service in SERVICES.items()}
    _save_runtime_state_cache(data)


def _restore_runtime_states() -> None:
    data = _load_runtime_state_cache()
    for name, service in SERVICES.items():
        service.load_persisted_state(data.get(name))


SERVICES: Dict[str, LiveMonitorService] = {
    name: LiveMonitorService(adapter) for name, adapter in ADAPTERS.items()
}

_restore_runtime_states()


def _resolve_platform(payload: dict) -> Optional[str]:
    p = (payload.get("platform") or "").strip().lower()
    if p in SERVICES:
        return p
    return None


def _resolve_platforms_from_payload(payload: dict, *, fallback_all: bool = False) -> List[str]:
    platforms = _normalize_platform_list(payload.get("platforms"))
    if platforms:
        return platforms
    single = _resolve_platform(payload)
    if single:
        return [single]
    if fallback_all:
        return list(PLATFORMS)
    return []


@app.route("/")
def index():
    return render_template(
        "index.html",
        platforms=[
            {
                "name": name,
                "label": ADAPTERS[name].label,
                "default_keywords": ", ".join(ADAPTERS[name].default_keywords),
            }
            for name in PLATFORMS
        ],
    )


@app.route("/api/start", methods=["POST"])
def api_start():
    payload = request.get_json(silent=True) or {}
    target_platforms = _resolve_platforms_from_payload(payload)
    if not target_platforms:
        return jsonify({"ok": False, "message": "请至少选择一个平台"}), 400
    try:
        cfg, selected_platforms = _build_monitor_config_from_payload(None, payload)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400

    for platform in PLATFORMS:
        if platform in selected_platforms:
            SERVICES[platform].start(cfg)
        else:
            SERVICES[platform].stop()
    return jsonify({
        "ok": True,
        "message": f"{'、'.join(ADAPTERS[name].label for name in selected_platforms)} 监控已启动",
        "platforms": selected_platforms,
        "config": cfg.__dict__,
    })


@app.route("/api/stop", methods=["POST"])
def api_stop():
    payload = request.get_json(silent=True) or {}
    target_platforms = _resolve_platforms_from_payload(payload, fallback_all=True)
    for platform in target_platforms:
        SERVICES[platform].stop()
    return jsonify({
        "ok": True,
        "platforms": target_platforms,
        "message": f"{'、'.join(ADAPTERS[name].label for name in target_platforms)} 监控已停止",
    })


@app.route("/api/state", methods=["GET"])
def api_state():
    for service in SERVICES.values():
        service.sanitize_for_api_state()
    platform_states = {name: SERVICES[name].state() for name in PLATFORMS}
    return jsonify({
        "ok": True,
        "data": {
            "platforms": platform_states,
            "platform_order": list(PLATFORMS),
            "labels": {name: ADAPTERS[name].label for name in PLATFORMS},
            "unified_config": _unified_monitor_config_summary(),
            "selected_platforms": [
                name for name in PLATFORMS if (platform_states.get(name, {}).get("summary") or {}).get("running")
            ] or list(_unified_monitor_config_summary().get("platforms") or []),
            "douyin_cookie": _douyin_cookie_summary(),
            "douyin_cookie_auto": DOUYIN_COOKIE_AUTO_STATE.snapshot(),
        },
    })


@app.route("/api/resume/options", methods=["GET"])
def api_resume_options():
    platform = (request.args.get("platform") or "").strip().lower()
    if platform not in SERVICES:
        return jsonify({"ok": False, "message": "缺少或非法的 platform 参数"}), 400
    return jsonify({
        "ok": True,
        "platform": platform,
        "data": {
            "files": _resume_export_options(platform),
        },
    })


@app.route("/api/resume/restore", methods=["POST"])
def api_resume_restore():
    payload = request.get_json(silent=True) or {}
    platform = _resolve_platform(payload)
    if not platform:
        return jsonify({"ok": False, "message": "缺少或非法的 platform 参数"}), 400
    file_name = str(payload.get("file_name") or "").strip()
    sheet_title = str(payload.get("sheet_title") or "").strip()
    if not file_name or not sheet_title:
        return jsonify({"ok": False, "message": "请选择导出文件和恢复节点"}), 400
    try:
        data = _restore_service_from_export(platform, file_name, sheet_title)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "message": f"恢复节点失败: {exc}"}), 500
    return jsonify({
        "ok": True,
        "platform": platform,
        "message": f"{ADAPTERS[platform].label} 已从节点恢复到第 {data['cycle_count']} 轮",
        "data": data,
    })


@app.route("/api/config/save", methods=["POST"])
def api_config_save():
    payload = request.get_json(silent=True) or {}
    try:
        data = _set_unified_monitor_config(payload)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({
        "ok": True,
        "platforms": list(data.get("platforms") or []),
        "message": "统一配置已保存",
        "data": data,
    })


@app.route("/api/config/unlock", methods=["POST"])
def api_config_unlock():
    saved = _unified_monitor_config_summary()
    if not saved.get("saved"):
        return jsonify({"ok": False, "message": "当前还没有已保存配置"}), 400
    data = _set_unified_monitor_lock(False)
    return jsonify({
        "ok": True,
        "platforms": list(data.get("platforms") or []),
        "message": "统一配置已解锁，可重新配置",
        "data": data,
    })


@app.route("/api/wecom/config", methods=["POST"])
def api_wecom_config():
    payload = request.get_json(silent=True) or {}
    platform = _resolve_platform(payload)
    if not platform:
        return jsonify({"ok": False, "message": "缺少或非法的 platform 参数"}), 400
    webhook_url = str(payload.get("webhook_url") or payload.get("wecom_webhook_url") or "").strip()
    auto_push = bool(payload.get("auto_push", payload.get("wecom_auto_push", True)))
    if webhook_url and not webhook_url.startswith("https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key="):
        return jsonify({"ok": False, "message": "企业微信 webhook 地址格式不正确"}), 400
    data = _set_wecom_config(platform, webhook_url=webhook_url, auto_push=auto_push)
    if webhook_url:
        message = f"{ADAPTERS[platform].label} 企业微信 webhook 已保存"
    else:
        message = f"{ADAPTERS[platform].label} 企业微信 webhook 已清空"
    return jsonify({"ok": True, "message": message, "data": data})


@app.route("/api/wecom/push", methods=["POST"])
def api_wecom_push():
    payload = request.get_json(silent=True) or {}
    target_platforms = _resolve_platforms_from_payload(payload)
    if not target_platforms:
        return jsonify({"ok": False, "message": "请至少选择一个平台"}), 400
    webhook_url = str(payload.get("webhook_url") or payload.get("wecom_webhook_url") or "").strip() or _unified_monitor_config_summary().get("wecom_webhook_url", "")
    if not webhook_url:
        return jsonify({"ok": False, "message": "请先配置企业微信 webhook"}), 400
    results = []
    all_ok = True
    for platform in target_platforms:
        ok, message = SERVICES[platform].push_current_to_wecom(webhook_url)
        all_ok = all_ok and ok
        results.append({"platform": platform, "ok": ok, "message": message})
    status = 200 if all_ok else 500
    return jsonify({
        "ok": all_ok,
        "message": "；".join(f"{ADAPTERS[item['platform']].label}: {item['message']}" for item in results),
        "platforms": target_platforms,
        "data": {
            "results": results,
            "services": {name: SERVICES[name].state() for name in target_platforms},
        },
    }), status


@app.route("/api/wecom/summary/push", methods=["POST"])
def api_wecom_summary_push():
    payload = request.get_json(silent=True) or {}
    target_platforms = _resolve_platforms_from_payload(payload)
    if not target_platforms:
        return jsonify({"ok": False, "message": "请至少选择一个平台"}), 400
    start_at = str(payload.get("start_at") or "").strip()
    end_at = str(payload.get("end_at") or "").strip()
    if not start_at or not end_at:
        return jsonify({"ok": False, "message": "请选择起始节点和结束节点"}), 400
    if not _parse_dt(start_at) or not _parse_dt(end_at):
        return jsonify({"ok": False, "message": "节点时间格式不正确"}), 400
    webhook_url = str(payload.get("webhook_url") or payload.get("wecom_webhook_url") or "").strip() or _unified_monitor_config_summary().get("wecom_webhook_url", "")
    if not webhook_url:
        return jsonify({"ok": False, "message": "请先配置企业微信 webhook"}), 400
    results = []
    all_ok = True
    for platform in target_platforms:
        ok, message = SERVICES[platform].push_summary_to_wecom(
            webhook_url,
            start_at=start_at,
            end_at=end_at,
        )
        all_ok = all_ok and ok
        results.append({"platform": platform, "ok": ok, "message": message})
    status = 200 if all_ok else 500
    return jsonify({
        "ok": all_ok,
        "message": "；".join(f"{ADAPTERS[item['platform']].label}: {item['message']}" for item in results),
        "platforms": target_platforms,
        "data": {"results": results},
    }), status


@app.route("/api/douyin/cookie", methods=["GET", "POST"])
def api_douyin_cookie():
    if request.method == "GET":
        return jsonify({"ok": True, "data": _douyin_cookie_summary()})

    payload = request.get_json(silent=True) or {}
    raw = payload.get("cookie") or payload.get("raw") or ""
    if not str(raw).strip():
        return jsonify({"ok": False, "message": "请填写抖音 Cookie 内容"}), 400
    try:
        _cookie_str, count = _save_douyin_cookie(str(raw))
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "message": f"保存失败: {exc}"}), 500
    return jsonify({
        "ok": True,
        "message": f"抖音 Cookie 已保存（共 {count} 个字段）",
        "data": _douyin_cookie_summary(),
    })


@app.route("/api/douyin/cookie/auto/status", methods=["GET"])
def api_douyin_cookie_auto_status():
    return jsonify({"ok": True, "data": DOUYIN_COOKIE_AUTO_STATE.snapshot()})


@app.route("/api/douyin/cookie/auto/start", methods=["POST"])
def api_douyin_cookie_auto_start():
    payload = request.get_json(silent=True) or {}
    keyword = str(payload.get("keyword") or "").strip()
    try:
        data = _start_cookie_browser_session(keyword)
    except Exception as exc:  # noqa: BLE001
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            DOUYIN_COOKIE_AUTO_STATE.last_error = str(exc)
        return jsonify({"ok": False, "message": f"启动自动获取 Cookie 失败: {exc}"}), 500
    return jsonify({
        "ok": True,
        "message": "自动化浏览器已打开，请扫码登录后点击“刷新获取 Cookie”",
        "data": data,
    })


@app.route("/api/douyin/cookie/auto/refresh", methods=["POST"])
def api_douyin_cookie_auto_refresh():
    try:
        auto_data = _refresh_cookie_from_auto_browser()
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    except Exception as exc:  # noqa: BLE001
        with DOUYIN_COOKIE_AUTO_STATE.lock:
            DOUYIN_COOKIE_AUTO_STATE.last_error = str(exc)
        return jsonify({"ok": False, "message": f"自动刷新 Cookie 失败: {exc}"}), 500
    return jsonify({
        "ok": True,
        "message": f"已刷新并保存最新抖音 Cookie（{auto_data.get('last_cookie_fields', 0)} 个字段）",
        "data": {
            "cookie": _douyin_cookie_summary(),
            "auto": auto_data,
        },
    })


@app.route("/api/exclude", methods=["POST"])
def api_exclude():
    payload = request.get_json(silent=True) or {}
    platform = _resolve_platform(payload)
    if not platform:
        return jsonify({"ok": False, "message": "缺少或非法的 platform 参数"}), 400
    action = payload.get("action") or "add"
    keys = payload.get("keys") or []
    svc = SERVICES[platform]
    if action == "clear":
        result = svc.clear_excluded()
        return jsonify({"ok": True, "data": result, "message": "已恢复全部被剔除直播间"})
    if not isinstance(keys, list) or not keys:
        return jsonify({"ok": False, "message": "请先选择要剔除的直播间"}), 400
    result = svc.exclude_rooms(keys)
    return jsonify({"ok": True, "data": result, "message": "已剔除选中直播间"})


@app.route("/api/export", methods=["POST"])
def api_export():
    payload = request.get_json(silent=True) or {}
    target_platforms = _resolve_platforms_from_payload(payload, fallback_all=True)
    export_mode = str(payload.get("mode") or "full").strip().lower()
    start_at = str(payload.get("start_at") or "").strip()
    end_at = str(payload.get("end_at") or "").strip()

    files: List[Tuple[str, Path]] = []
    for name in target_platforms:
        svc = SERVICES[name]
        if export_mode == "current":
            path = svc.export_current_node_xlsx()
        elif export_mode == "summary":
            path = svc.export_summary_range_xlsx(start_at=start_at, end_at=end_at)
        else:
            path = svc.export_single_xlsx()
        files.append((name, path))

    if len(files) == 1:
        _, path = files[0]
        return send_file(path, as_attachment=True, download_name=path.name)

    buf = io.BytesIO()
    zip_name = f"直播监控导出_合集_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, path in files:
            zf.write(path, arcname=path.name)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=zip_name,
        mimetype="application/zip",
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5056, debug=False, use_reloader=False, threaded=True)
